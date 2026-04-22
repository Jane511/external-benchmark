"""Shared fixtures for ingestion tests.

Primary contribution: `apra_xlsx_fixture` builds a synthetic XLSX whose
shape matches what `apra_adi.ApraAdiScraper` expects (as declared in
`ingestion/config/sources.yaml`). Generating it at test time avoids
committing binary files and keeps the fixture aligned with the config.
"""
from __future__ import annotations

from datetime import date
from pathlib import Path

import openpyxl
import pytest
import yaml


@pytest.fixture()
def sources_config() -> dict:
    path = (
        Path(__file__).resolve().parent.parent.parent
        / "ingestion" / "config" / "sources.yaml"
    )
    with open(path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


@pytest.fixture()
def apra_xlsx_fixture(tmp_path: Path) -> Path:
    """Write a synthetic APRA ADI XLSX to tmp_path and return the path.

    Values are plausible 2025 Q3 Australian readings:
      - Residential 90+DPD ~ 1.2%
      - Commercial  90+DPD ~ 1.8%
      - Corporate   90+DPD ~ 2.5%
    """
    path = tmp_path / "apra_adi_fixture.xlsx"
    wb = openpyxl.Workbook()

    # Default sheet is "Sheet" — rename to Asset Quality
    ws = wb.active
    ws.title = "Asset Quality"
    ws.append(["Period", "Category", "90DPD_Rate", "NPL_Rate"])
    ws.append([date(2025, 9, 30), "Residential", 0.012, 0.008])
    ws.append([date(2025, 9, 30), "Commercial", 0.018, 0.012])
    ws.append([date(2025, 9, 30), "Corporate", 0.025, 0.015])

    # Second sheet — Property Exposures (QPEX shape)
    qpex = wb.create_sheet("Property Exposures")
    qpex.append(["Period", "Category", "Non_Performing_Ratio"])
    qpex.append([date(2025, 9, 30), "Residential", 0.009])
    qpex.append([date(2025, 9, 30), "Commercial", 0.014])

    wb.save(path)
    return path


@pytest.fixture()
def cba_pillar3_xlsx(tmp_path: Path) -> Path:
    """Synthetic CBA Pillar 3 Excel companion matching `cba_pillar3` in sources.yaml.

    Shape mirrors the expected workbook layout:
      Sheet "IRB Credit Risk"     : Portfolio | Exposure_EAD_Mn | PD | LGD
      Sheet "Specialised Lending" : Grade | PD
    Values are plausible CBA H1 FY2025 readings.
    """
    path = tmp_path / "sample_cba_pillar3.xlsx"
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "IRB Credit Risk"
    ws.append(["Portfolio", "Exposure_EAD_Mn", "PD", "LGD"])
    ws.append(["Residential Mortgage", 500000, 0.0072, 0.22])
    ws.append(["CRE Investment",       80000,  0.0250, 0.35])
    ws.append(["Corporate SME",        120000, 0.0280, 0.45])

    sl = wb.create_sheet("Specialised Lending")
    sl.append(["Grade", "PD"])
    sl.append(["Strong",       0.004])
    sl.append(["Good",         0.008])
    sl.append(["Satisfactory", 0.028])
    sl.append(["Weak",         0.080])

    wb.save(path)
    return path


@pytest.fixture()
def apra_xlsx_bad_values(tmp_path: Path) -> Path:
    """Same shape as apra_xlsx_fixture but one row carries an out-of-range value.

    Used to verify ApraAdiScraper.validate() drops the bad row.
    """
    path = tmp_path / "apra_adi_bad.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Asset Quality"
    ws.append(["Period", "Category", "90DPD_Rate", "NPL_Rate"])
    ws.append([date(2025, 9, 30), "Residential", 0.012, 0.008])
    ws.append([date(2025, 9, 30), "Commercial", 0.95, 0.012])   # 95% 90DPD — implausible
    wb.save(path)
    return path
