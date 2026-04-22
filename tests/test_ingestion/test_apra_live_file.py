"""Integration tests: ApraAdiScraper routing between canonical and adapter paths."""
from __future__ import annotations

from datetime import date
from pathlib import Path
from unittest.mock import patch

import openpyxl
import pytest

from ingestion.adapters.apra_performance_adapter import SYNTHETIC_ASSET_CLASS
from ingestion.apra_adi import ApraAdiScraper


@pytest.fixture()
def live_performance_workbook(tmp_path: Path) -> Path:
    """Minimal live-shaped workbook: Tab 2d with two APS-220 quarters."""
    path = tmp_path / "live_perf.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "Cover"

    ws = wb.create_sheet("Tab 2d")
    ws.cell(row=1, column=1, value="Table 2d  Banks' asset quality")
    ws.cell(row=3, column=2, value="Quarter end")
    # Date header row (row 4 = index 3): populate 12 columns so the adapter's
    # ">= 10 date cells" threshold passes. Values before Mar 2022 are ignored
    # by the APS-220 floor anyway.
    dates = [
        "Mar 2021", "Jun 2021", "Sep 2021", "Dec 2021",
        "Mar 2022", "Jun 2022", "Sep 2022", "Dec 2022",
        "Mar 2023", "Jun 2023", "Sep 2023", "Dec 2023",
    ]
    for i, d in enumerate(dates):
        ws.cell(row=4, column=2 + i, value=d)

    # NPL ratio row (row 54 = index 53); only Mar-2022+ populated.
    ws.cell(row=54, column=1, value="Non-performing to loans and advances")
    npl = [None, None, None, None, 0.009, 0.010, 0.010, 0.011, 0.011, 0.011, 0.012, 0.012]
    for i, v in enumerate(npl):
        if v is not None:
            ws.cell(row=54, column=2 + i, value=v)

    # 90DPD row (row 55 = index 54)
    ws.cell(row=55, column=1,
            value="  of which >= 90 days past due to loans and advances")
    dpd = [None, None, None, None, 0.005, 0.006, 0.006, 0.007, 0.007, 0.007, 0.006, 0.007]
    for i, v in enumerate(dpd):
        if v is not None:
            ws.cell(row=55, column=2 + i, value=v)

    wb.save(path)
    return path


def _perf_config() -> dict:
    """Minimal config matching sources.yaml apra_adi_performance."""
    return {
        "source_name": "apra_adi_performance",
        "scraper": "apra_adi",
        "url": "https://www.apra.gov.au/",
        "frequency_days": 120,
        "quality_score": "HIGH",
        "sheet": "Asset Quality",
        "period_column": "Period",
        "category_column": "Category",
        "file_name": "ADI Performance",
        "metrics": [
            {"column": "90DPD_Rate", "metric": "90dpd",
             "data_type": "impaired_ratio", "unit": "ratio",
             "validation_range": [0.0, 0.20]},
            {"column": "NPL_Rate", "metric": "npl",
             "data_type": "impaired_ratio", "unit": "ratio",
             "validation_range": [0.0, 0.30]},
        ],
    }


def test_scraper_routes_live_file_through_adapter(
    live_performance_workbook: Path,
) -> None:
    scraper = ApraAdiScraper(
        source_path=live_performance_workbook,
        config=_perf_config(),
        retrieval_date=date(2026, 4, 21),
    )

    with patch(
        "ingestion.apra_adi.ApraPerformanceAdapter.normalise",
        wraps=__import__(
            "ingestion.adapters.apra_performance_adapter", fromlist=["ApraPerformanceAdapter"],
        ).ApraPerformanceAdapter().normalise,
    ) as spy:
        points = scraper.scrape()

    assert spy.call_count == 1, "adapter must be invoked for live-shaped file"
    assert points, "adapter path produced zero points"
    for p in points:
        assert p.publisher == "APRA"
        assert p.asset_class_raw == SYNTHETIC_ASSET_CLASS
        assert p.quality_indicators.get("adapter") == "ApraPerformanceAdapter"
        assert p.quality_indicators.get("sector") in {
            "all_adis", "banks", "major_banks",
        }
        # period_code must be an annual/quarterly slug that survives into source_id.
        assert p.metadata["period_code"].endswith(("Q1", "Q2", "Q3", "Q4"))


def test_scraper_bypasses_adapter_for_canonical_fixture(
    apra_xlsx_fixture: Path,
) -> None:
    """Fixture has 'Asset Quality' sheet → direct-read path, adapter untouched."""
    scraper = ApraAdiScraper(
        source_path=apra_xlsx_fixture,
        config=_perf_config(),
        retrieval_date=date(2025, 10, 1),
    )

    with patch(
        "ingestion.apra_adi.ApraPerformanceAdapter",
    ) as adapter_cls:
        points = scraper.scrape()

    adapter_cls.assert_not_called()
    assert points, "canonical path produced zero points"
    # Canonical fixture encodes residential/commercial/corporate categories,
    # so asset_class_raw should NOT be the synthetic sector-total label.
    assert all(p.asset_class_raw != SYNTHETIC_ASSET_CLASS for p in points)


def test_qpex_live_file_routes_to_qpex_adapter(tmp_path: Path) -> None:
    """Live QPEX-shaped workbook must flow through ApraQpexAdapter and yield
    residential_mortgage / commercial_property_investment ScrapedDataPoints."""
    # Re-use the QPEX adapter test's builder for a realistic synthetic file.
    from tests.test_ingestion.adapters.test_apra_qpex_adapter import (
        live_qpex_workbook as _build_live_qpex,  # pytest fixture function
    )
    # pytest fixture functions aren't directly callable — inline the build.
    import openpyxl
    from tests.test_ingestion.adapters.test_apra_qpex_adapter import (
        _commercial_sheet, _residential_sheet,
    )

    path = tmp_path / "qpex_live.xlsx"
    wb = openpyxl.Workbook(); wb.active.title = "Cover"
    comm_total   = [349_355.4 + 4_000 * i for i in range(16)]
    comm_nonperf = [1_855.3 + 90 * i for i in range(16)]
    res_total    = [1_700_000.0 + 25_000 * i for i in range(16)]
    res_nonperf  = [11_500.0 + 550 * i for i in range(16)]
    _commercial_sheet(
        wb.create_sheet("Tab 2a"),
        title="Table 2a  Banks' commercial property exposures",
        total_row_idx=25, impaired_row_idx=28, nonperf_row_idx=42,
        total_values=comm_total, nonperf_values=comm_nonperf,
    )
    _residential_sheet(
        wb.create_sheet("Tab 2b"),
        title="Table 2b Banks' residential property exposures",
        total_values=res_total, nonperf_values=res_nonperf,
    )
    wb.save(path)

    qpex_config = _perf_config()
    qpex_config["source_name"] = "apra_qpex"
    qpex_config["file_name"] = "Property Exposures"
    qpex_config["sheet"] = "Property Exposures"

    scraper = ApraAdiScraper(
        source_path=path,
        config=qpex_config,
        retrieval_date=date(2026, 4, 21),
    )

    with patch(
        "ingestion.apra_adi.ApraQpexAdapter.normalise",
        wraps=__import__(
            "ingestion.adapters.apra_qpex_adapter",
            fromlist=["ApraQpexAdapter"],
        ).ApraQpexAdapter().normalise,
    ) as spy:
        points = scraper.scrape()

    assert spy.call_count == 1, "QPEX adapter must be invoked for live QPEX"
    assert points, "QPEX adapter path produced zero points"
    assert {p.asset_class_raw for p in points} == {
        "commercial_property_investment",
        "residential_mortgage",
    }
    # All QPEX points must carry numerator/denominator for MRC audit.
    for p in points:
        qi = p.quality_indicators
        assert qi["adapter"] == "ApraQpexAdapter"
        assert qi["numerator_value"] > 0
        assert qi["denominator_value"] > 0
        assert "arithmetic" in qi


def test_qpex_routing_does_not_invoke_performance_adapter(tmp_path: Path) -> None:
    """QPEX live file must NOT invoke the Performance adapter."""
    from tests.test_ingestion.adapters.test_apra_qpex_adapter import (
        _residential_sheet,
    )
    import openpyxl

    path = tmp_path / "qpex_only.xlsx"
    wb = openpyxl.Workbook(); wb.active.title = "Cover"
    _residential_sheet(
        wb.create_sheet("Tab 2b"),
        title="Table 2b Banks' residential property exposures",
    )
    wb.save(path)

    qpex_config = _perf_config()
    qpex_config["source_name"] = "apra_qpex"
    qpex_config["file_name"] = "Property Exposures"
    qpex_config["sheet"] = "Property Exposures"

    scraper = ApraAdiScraper(
        source_path=path, config=qpex_config,
        retrieval_date=date(2026, 4, 21),
    )

    with patch("ingestion.apra_adi.ApraPerformanceAdapter") as perf_cls:
        scraper.scrape()
    perf_cls.assert_not_called()


def test_performance_live_file_still_bypasses_qpex_adapter(
    live_performance_workbook: Path,
) -> None:
    """Path A regression: Performance live file must NOT touch the QPEX adapter."""
    scraper = ApraAdiScraper(
        source_path=live_performance_workbook,
        config=_perf_config(),
        retrieval_date=date(2026, 4, 21),
    )
    with patch("ingestion.apra_adi.ApraQpexAdapter") as qpex_cls:
        points = scraper.scrape()
    qpex_cls.assert_not_called()
    assert points, "Performance path still produced points"
