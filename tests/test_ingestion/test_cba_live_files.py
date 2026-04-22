"""Integration tests: CBAScraper routes between three input paths.

- .xlsx with canonical sheets → direct-read path (existing fixture tests)
- .xlsx live quarterly APS 330  → CbaPillar3QuarterlyAdapter (Option A)
- .pdf live annual              → CbaPillar3PdfAdapter (Option B)

No real PDFs are committed; the PDF routing test patches the adapter's
``normalise`` to return a small canned DataFrame.
"""
from __future__ import annotations

from datetime import date
from pathlib import Path
from unittest.mock import patch

import openpyxl
import pandas as pd
import pytest

from ingestion.pillar3.cba import CBAScraper


def _cba_config() -> dict:
    return {
        "source_name": "cba_pillar3",
        "scraper": "pillar3_cba",
        "url": "https://www.commbank.com.au/",
        "frequency_days": 180,
        "quality_score": "HIGH",
        "irb_sheet": "IRB Credit Risk",
        "portfolio_column": "Portfolio",
        "ead_column": "Exposure_EAD_Mn",
        "pd_column": "PD",
        "lgd_column": "LGD",
        "slotting_sheet": "Specialised Lending",
        "grade_column": "Grade",
        "slotting_pd_column": "PD",
    }


def test_canonical_xlsx_bypasses_both_adapters(cba_pillar3_xlsx: Path) -> None:
    """Fixture with 'IRB Credit Risk' + 'Specialised Lending' sheets must
    flow through the direct-read path; neither adapter is invoked."""
    scraper = CBAScraper(
        source_path=cba_pillar3_xlsx, config=_cba_config(),
        reporting_date=date(2025, 6, 30),
    )
    with patch("ingestion.pillar3.cba.CbaPillar3PdfAdapter") as pdf_cls, \
         patch("ingestion.pillar3.cba.CbaPillar3QuarterlyAdapter") as qx_cls:
        points = scraper.scrape()
    pdf_cls.assert_not_called()
    qx_cls.assert_not_called()
    assert points, "Canonical fixture produced zero points"


def test_live_quarterly_xlsx_routes_to_quarterly_adapter(tmp_path: Path) -> None:
    """Workbook lacking 'IRB Credit Risk' goes through the quarterly adapter."""
    path = tmp_path / "quarterly.xlsx"
    wb = openpyxl.Workbook(); wb.active.title = "Cover"
    ead = wb.create_sheet("EAD & CRWA")
    ead.cell(row=9, column=1, value="Residential mortgage")
    ead.cell(row=9, column=2, value=100000)
    crb = wb.create_sheet("CRB(f)")
    crb.cell(row=50, column=1, value="Residential Mortgage")
    crb.cell(row=50, column=2, value=900)
    wb.save(path)

    scraper = CBAScraper(
        source_path=path, config=_cba_config(),
        reporting_date=date(2025, 9, 30),
    )
    with patch("ingestion.pillar3.cba.CbaPillar3PdfAdapter") as pdf_cls:
        points = scraper.scrape()
    pdf_cls.assert_not_called()

    assert points, "Quarterly XLSX adapter produced zero points"
    assert all(p.publisher == "CBA" for p in points)
    assert all(p.quality_indicators.get("adapter") == "CbaPillar3QuarterlyAdapter"
               for p in points)
    # Arithmetic audit trail present on every quarterly point.
    for p in points:
        qi = p.quality_indicators
        assert qi["numerator_value"] > 0
        assert qi["denominator_value"] > 0
        assert "arithmetic" in qi


def test_live_pdf_routes_to_pdf_adapter(tmp_path: Path) -> None:
    """A .pdf path must flow through the PDF adapter; its normalise() output
    becomes ScrapedDataPoints tagged with the pd_band provenance."""
    # Create a real (but empty-content) PDF file — the scraper dispatches
    # on extension; we patch the adapter's normalise() so pdfplumber is
    # not actually invoked.
    pdf_path = tmp_path / "cba_annual.pdf"
    pdf_path.write_bytes(b"%PDF-1.4\n%%EOF\n")

    canned = pd.DataFrame([
        {
            "asset_class": "residential_mortgage", "metric_name": "pd",
            "value": 0.0008, "as_of_date": date(2025, 6, 30),
            "period_code": "FY2025", "value_basis": "exposure_weighted",
            "source_table": "CR6", "source_page": 55, "pd_band": "0.00 to <0.15",
        },
        {
            "asset_class": "residential_mortgage", "metric_name": "lgd",
            "value": 0.14, "as_of_date": date(2025, 6, 30),
            "period_code": "FY2025", "value_basis": "exposure_weighted",
            "source_table": "CR6", "source_page": 55, "pd_band": "0.00 to <0.15",
        },
        {
            "asset_class": "development_strong", "metric_name": "risk_weight",
            "value": 0.70, "as_of_date": date(2025, 6, 30),
            "period_code": "FY2025", "value_basis": "supervisory_prescribed",
            "source_table": "CR10", "source_page": 48, "pd_band": "all",
        },
    ])

    scraper = CBAScraper(
        source_path=pdf_path, config=_cba_config(),
        reporting_date=date(2025, 6, 30),
    )
    with patch(
        "ingestion.pillar3.cba.CbaPillar3PdfAdapter.normalise",
        return_value=canned,
    ) as spy, \
         patch("ingestion.pillar3.cba.CbaPillar3QuarterlyAdapter") as qx_cls:
        points = scraper.scrape()

    assert spy.call_count == 1
    qx_cls.assert_not_called()
    assert len(points) == 3
    by_metric = {(p.asset_class_raw, p.metadata["data_type_hint"]) for p in points}
    assert ("residential_mortgage", "pd") in by_metric
    assert ("residential_mortgage", "lgd") in by_metric
    assert ("development_strong", "supervisory") in by_metric
    # Provenance is stamped on every point.
    for p in points:
        qi = p.quality_indicators
        assert qi["adapter"] == "CbaPillar3PdfAdapter"
        assert qi["source_table"] in {"CR6", "CR10"}
        assert "pd_band" in qi


def test_unknown_extension_logs_and_returns_empty(tmp_path: Path, caplog) -> None:
    import logging
    caplog.set_level(logging.WARNING)
    path = tmp_path / "strange.txt"
    path.write_text("not a real file")
    scraper = CBAScraper(source_path=path, config=_cba_config())
    assert scraper.scrape() == []
    assert any("unexpected extension" in r.message for r in caplog.records)
