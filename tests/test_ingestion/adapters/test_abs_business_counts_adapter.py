"""Tests for the ABS Counts of Australian Businesses (cat. 8165) adapter."""
from __future__ import annotations

from datetime import date
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from ingestion.adapters.abs_business_counts_adapter import (
    ANZSIC_DIVISION_CODES,
    AbsBusinessCountsAdapter,
)


def _build_abs_workbook(path: Path, *, include_table1: bool = True) -> None:
    """Synthesise a minimal 8165 DC01-shaped workbook (2 FYs × 3 divisions)."""
    wb = openpyxl.Workbook()
    wb.active.title = "Contents"
    if include_table1:
        ws = wb.create_sheet("Table 1")
        # Header band — rows 0-5 carry title/metadata that the adapter ignores.
        ws.cell(row=1, column=1, value="Australian Bureau of Statistics")
        ws.cell(row=4, column=1, value="Table 1 Businesses by Industry Division")
        ws.cell(row=5, column=2, value="Operating at start of financial year")
        ws.cell(row=5, column=9, value="Operating at end of financial year")
        ws.cell(row=6, column=2, value="no.")

        # FY 2023-24 panel (FY end = 2024-06-30). Openpyxl rows are 1-indexed
        # so row=7 corresponds to the adapter's index 6.
        ws.cell(row=7, column=1, value="2023-24")
        ws.cell(row=8, column=1, value="Agriculture, Forestry and Fishing")
        ws.cell(row=8, column=9, value=175000)
        ws.cell(row=9, column=1, value="Construction")
        ws.cell(row=9, column=9, value=440000)
        ws.cell(row=10, column=1, value="Retail Trade")
        ws.cell(row=10, column=9, value=158000)
        ws.cell(row=11, column=1, value="Currently Unknown")
        ws.cell(row=11, column=9, value=2500)      # must be excluded
        ws.cell(row=12, column=1, value="All Industries")
        ws.cell(row=12, column=9, value=2_600_000)  # must be excluded

        # FY 2024-25 panel (FY end = 2025-06-30)
        ws.cell(row=13, column=1, value="2024-25")
        ws.cell(row=14, column=1, value="Agriculture, Forestry and Fishing")
        ws.cell(row=14, column=9, value=176000)
        ws.cell(row=15, column=1, value="Construction")
        ws.cell(row=15, column=9, value=450000)
        ws.cell(row=16, column=1, value="Retail Trade")
        ws.cell(row=16, column=9, value=160000)
        # An implausibly small count — must be dropped by plausibility filter.
        ws.cell(row=17, column=1, value="Mining")
        ws.cell(row=17, column=9, value=50)
    wb.save(path)


@pytest.fixture()
def abs_workbook(tmp_path: Path) -> Path:
    p = tmp_path / "8165DC01.xlsx"
    _build_abs_workbook(p)
    return p


def test_adapter_finds_table_1_and_walks_panels(abs_workbook: Path) -> None:
    df = AbsBusinessCountsAdapter().normalise(abs_workbook)
    # 3 divisions × 2 FYs = 6 rows (Mining implausible dropped; "Currently
    # Unknown" and "All Industries" excluded).
    assert len(df) == 6
    assert set(df["fiscal_year"].unique()) == {"FY2024", "FY2025"}


def test_adapter_maps_anzsic_division_codes(abs_workbook: Path) -> None:
    df = AbsBusinessCountsAdapter().normalise(abs_workbook)
    codes = df[["industry", "anzsic_division_code"]].drop_duplicates()
    for _, row in codes.iterrows():
        expected = ANZSIC_DIVISION_CODES[row["industry"].lower()]
        assert row["anzsic_division_code"] == expected


def test_adapter_excludes_sweep_and_total_rows(abs_workbook: Path) -> None:
    df = AbsBusinessCountsAdapter().normalise(abs_workbook)
    industries = set(df["industry"].str.lower())
    assert "currently unknown" not in industries
    assert "all industries" not in industries


def test_adapter_drops_implausible_counts(abs_workbook: Path, caplog) -> None:
    import logging
    caplog.set_level(logging.WARNING)
    df = AbsBusinessCountsAdapter().normalise(abs_workbook)
    assert "Mining" not in set(df["industry"])
    assert any("implausible business_count" in r.message for r in caplog.records)


def test_adapter_maps_fy_label_to_end_date(abs_workbook: Path) -> None:
    df = AbsBusinessCountsAdapter().normalise(abs_workbook)
    fy2024 = df[df["fiscal_year"] == "FY2024"].iloc[0]
    assert fy2024["as_of_date"] == date(2024, 6, 30)
    fy2025 = df[df["fiscal_year"] == "FY2025"].iloc[0]
    assert fy2025["as_of_date"] == date(2025, 6, 30)


def test_adapter_produces_canonical_columns(abs_workbook: Path) -> None:
    adapter = AbsBusinessCountsAdapter()
    df = adapter.normalise(abs_workbook)
    for col in adapter.canonical_columns:
        assert col in df.columns


def test_adapter_handles_missing_table_1(tmp_path: Path, caplog) -> None:
    import logging
    caplog.set_level(logging.WARNING)
    p = tmp_path / "broken.xlsx"
    _build_abs_workbook(p, include_table1=False)
    df = AbsBusinessCountsAdapter().normalise(p)
    assert df.empty
    assert any("not in workbook" in r.message for r in caplog.records)


def test_validate_output_rejects_missing_column() -> None:
    adapter = AbsBusinessCountsAdapter()
    with pytest.raises(ValueError, match="missing required columns"):
        adapter.validate_output(pd.DataFrame({"industry": ["x"]}))
