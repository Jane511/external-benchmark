"""Tests for the live-APRA QPEX adapter (Path B)."""
from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from ingestion.adapters.apra_qpex_adapter import ApraQpexAdapter


# ---------------------------------------------------------------------------
# Synthetic workbook builder
# ---------------------------------------------------------------------------

_QUARTERLY_DATES = [
    datetime(2022, 3, 31), datetime(2022, 6, 30), datetime(2022, 9, 30),
    datetime(2022, 12, 31),
    datetime(2023, 3, 31), datetime(2023, 6, 30), datetime(2023, 9, 30),
    datetime(2023, 12, 31),
    datetime(2024, 3, 31), datetime(2024, 6, 30), datetime(2024, 9, 30),
    datetime(2024, 12, 31),
    datetime(2025, 3, 31), datetime(2025, 6, 30), datetime(2025, 9, 30),
    datetime(2025, 12, 31),
]


def _commercial_sheet(
    ws,
    *,
    title: str,
    total_row_idx: int,
    impaired_row_idx: int,
    nonperf_row_idx: int,
    total_values: list[float | None],
    nonperf_values: list[float | None],
) -> None:
    """Populate a QPEX commercial sheet (Tab ?a) with just the rows we need."""
    ws.cell(row=1, column=1, value=title)
    ws.cell(row=2, column=1, value=" ($ million, consolidated group)")
    ws.cell(row=3, column=2, value="Quarter end")
    for i, d in enumerate(_QUARTERLY_DATES):
        ws.cell(row=4, column=2 + i, value=d)

    # Total commercial property exposures row
    ws.cell(row=total_row_idx + 1, column=1, value="Total commercial property exposures")
    for i, v in enumerate(total_values):
        if v is not None:
            ws.cell(row=total_row_idx + 1, column=2 + i, value=v)
    # "of which" sub-row immediately after — must NOT be matched as total
    ws.cell(
        row=total_row_idx + 2, column=1,
        value="of which: Exposures in Australia",
    )

    # Legacy impaired row (not matched by numerator — present for realism)
    ws.cell(row=impaired_row_idx + 1, column=1,
            value="Impaired Commercial property exposures")

    # Non-performing (APS 220) — matched as numerator
    ws.cell(row=nonperf_row_idx + 1, column=1,
            value="Non-performing commercial property exposures")
    for i, v in enumerate(nonperf_values):
        if v is not None:
            ws.cell(row=nonperf_row_idx + 1, column=2 + i, value=v)
    # "of which" sub-row immediately after
    ws.cell(row=nonperf_row_idx + 2, column=1,
            value="of which: Exposures in Australia")


def _residential_sheet(
    ws,
    *,
    title: str,
    total_row_idx: int = 15,
    nonperf_row_idx: int = 68,
    total_values: list[float | None] | None = None,
    nonperf_values: list[float | None] | None = None,
) -> None:
    """Populate a QPEX residential sheet (Tab ?b)."""
    if total_values is None:
        total_values = [1_700_000.0 + 20_000 * i for i in range(len(_QUARTERLY_DATES))]
    if nonperf_values is None:
        nonperf_values = [12_000.0 + 500 * i for i in range(len(_QUARTERLY_DATES))]

    ws.cell(row=1, column=1, value=title)
    ws.cell(row=2, column=1, value="($ million, Level 2)")
    ws.cell(row=3, column=2, value="Quarter end")
    for i, d in enumerate(_QUARTERLY_DATES):
        ws.cell(row=4, column=2 + i, value=d)

    # Credit-limits total (row 6) — must NOT be matched (excluded by "limits"
    # in forbidden patterns).
    ws.cell(row=7, column=1,
            value="Total credit limits (including redrawable amounts)")

    # Denominator: Total credit oustanding (sic)
    ws.cell(row=total_row_idx + 1, column=1, value="Total credit oustanding")
    for i, v in enumerate(total_values):
        if v is not None:
            ws.cell(row=total_row_idx + 1, column=2 + i, value=v)

    # "Term loans - by purpose" breakdown (row 16) — must NOT match
    ws.cell(row=total_row_idx + 2, column=1, value="Term loans - by purpose")

    # Numerator: Non-performing loans (row 68)
    ws.cell(row=nonperf_row_idx + 1, column=1, value="Non-performing loans")
    for i, v in enumerate(nonperf_values):
        if v is not None:
            ws.cell(row=nonperf_row_idx + 1, column=2 + i, value=v)

    # Breakdown rows that must NOT be matched as numerator
    ws.cell(row=nonperf_row_idx + 2, column=1, value="Term loans")
    ws.cell(row=nonperf_row_idx + 3, column=1, value="Owner-occupied")


@pytest.fixture()
def live_qpex_workbook(tmp_path: Path) -> Path:
    """Synthetic QPEX with all 6 sheets; row indices mirror the live file."""
    path = tmp_path / "live_qpex.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "Cover"
    wb.create_sheet("Contents")

    # 16 quarters of plausible commercial and residential series.
    comm_total   = [349_355.4 + 4_000 * i for i in range(16)]
    comm_nonperf = [1_855.3 + 90 * i for i in range(16)]
    res_total    = [1_700_000.0 + 25_000 * i for i in range(16)]
    res_nonperf  = [11_500.0 + 550 * i for i in range(16)]

    # Commercial — Tab 1a uses rows 26/29/43 (offset by 1), Tab 2a/4a use 25/28/42.
    _commercial_sheet(
        wb.create_sheet("Tab 1a"),
        title="Table 1a  All ADIs' commercial property exposures",
        total_row_idx=26, impaired_row_idx=29, nonperf_row_idx=43,
        total_values=comm_total, nonperf_values=comm_nonperf,
    )
    _commercial_sheet(
        wb.create_sheet("Tab 2a"),
        title="Table 2a  Banks' commercial property exposures",
        total_row_idx=25, impaired_row_idx=28, nonperf_row_idx=42,
        total_values=comm_total, nonperf_values=comm_nonperf,
    )
    _commercial_sheet(
        wb.create_sheet("Tab 4a"),
        title="Table 4a  Major banks' commercial property exposures",
        total_row_idx=25, impaired_row_idx=28, nonperf_row_idx=42,
        total_values=comm_total, nonperf_values=comm_nonperf,
    )

    # Residential — Tab 1b/2b/4b all use the same rows 15 / 68.
    for label, sheet_name in (("ADIs'", "Tab 1b"),
                              ("Banks'", "Tab 2b"),
                              ("Major banks'", "Tab 4b")):
        _residential_sheet(
            wb.create_sheet(sheet_name),
            title=f"Table  {label} residential property exposures",
            total_values=res_total, nonperf_values=res_nonperf,
        )

    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


def test_adapter_finds_all_six_sector_sheets(live_qpex_workbook: Path) -> None:
    df = ApraQpexAdapter().normalise(live_qpex_workbook)
    pairs = df.groupby(["institution_sector", "asset_class"]).size()
    assert set(pairs.index) == {
        ("all_adis", "commercial_property_investment"),
        ("all_adis", "residential_mortgage"),
        ("banks", "commercial_property_investment"),
        ("banks", "residential_mortgage"),
        ("major_banks", "commercial_property_investment"),
        ("major_banks", "residential_mortgage"),
    }


def test_adapter_yields_expected_row_count(live_qpex_workbook: Path) -> None:
    df = ApraQpexAdapter().normalise(live_qpex_workbook)
    # 6 sheets × 16 quarters = 96 rows.
    assert len(df) == 96


def test_adapter_computes_npl_ratio_correctly(live_qpex_workbook: Path) -> None:
    df = ApraQpexAdapter().normalise(live_qpex_workbook)
    first = df[df["institution_sector"] == "banks"].iloc[0]
    expected_ratio = first["numerator_value"] / first["denominator_value"]
    assert first["value"] == pytest.approx(expected_ratio, rel=1e-9)


def test_adapter_handles_zero_denominator(tmp_path: Path) -> None:
    path = tmp_path / "zero_den.xlsx"
    wb = openpyxl.Workbook(); wb.active.title = "Cover"
    _residential_sheet(
        wb.create_sheet("Tab 2b"),
        title="Table 2b Banks' residential property exposures",
        total_values=[0.0] * 16,
        nonperf_values=[100.0] * 16,
    )
    wb.save(path)
    df = ApraQpexAdapter().normalise(path)
    assert df.empty  # division-by-zero rows dropped, no ZeroDivisionError


def test_adapter_filters_implausible_ratios(tmp_path: Path, caplog) -> None:
    import logging
    caplog.set_level(logging.WARNING)
    path = tmp_path / "implausible.xlsx"
    wb = openpyxl.Workbook(); wb.active.title = "Cover"
    # 50% NPL is 5× the PLAUSIBILITY upper bound of 10% — must be filtered.
    bad_nonperf = [11_500.0] * 15 + [850_000.0]     # last quarter = 50% NPL
    _residential_sheet(
        wb.create_sheet("Tab 2b"),
        title="Table 2b Banks' residential property exposures",
        total_values=[1_700_000.0] * 16,
        nonperf_values=bad_nonperf,
    )
    wb.save(path)
    df = ApraQpexAdapter().normalise(path)
    assert (df["value"] <= 0.10).all()
    assert len(df) == 15
    assert any("implausible" in r.message.lower() for r in caplog.records)


def test_adapter_preserves_numerator_and_denominator(live_qpex_workbook: Path) -> None:
    df = ApraQpexAdapter().normalise(live_qpex_workbook)
    assert "numerator_value" in df.columns
    assert "denominator_value" in df.columns
    assert (df["numerator_value"] > 0).all()
    assert (df["denominator_value"] > 0).all()


def test_adapter_maps_asset_class_correctly(live_qpex_workbook: Path) -> None:
    df = ApraQpexAdapter().normalise(live_qpex_workbook)
    # Tab ?a → commercial; Tab ?b → residential
    tabs_a = df[df["_source_sheet"].str.endswith("a")]
    tabs_b = df[df["_source_sheet"].str.endswith("b")]
    assert (tabs_a["asset_class"] == "commercial_property_investment").all()
    assert (tabs_b["asset_class"] == "residential_mortgage").all()


def test_adapter_snaps_dates_to_quarter_end(live_qpex_workbook: Path) -> None:
    df = ApraQpexAdapter().normalise(live_qpex_workbook)
    expected_days = {3: 31, 6: 30, 9: 30, 12: 31}
    for d in df["as_of_date"]:
        assert d.day == expected_days[d.month]
        assert d.month in (3, 6, 9, 12)


def test_adapter_produces_canonical_columns(live_qpex_workbook: Path) -> None:
    adapter = ApraQpexAdapter()
    df = adapter.normalise(live_qpex_workbook)
    for col in adapter.canonical_columns:
        assert col in df.columns


def test_adapter_handles_missing_sector_sheet(
    live_qpex_workbook: Path, caplog,
) -> None:
    """Drop Tab 4a mid-file; adapter should warn + emit rows from others."""
    import logging
    caplog.set_level(logging.WARNING)
    wb = openpyxl.load_workbook(live_qpex_workbook)
    del wb["Tab 4a"]
    alt = live_qpex_workbook.with_name("no_tab4a.xlsx")
    wb.save(alt)
    df = ApraQpexAdapter().normalise(alt)
    sectors_assets = set(zip(df["institution_sector"], df["asset_class"]))
    assert ("major_banks", "commercial_property_investment") not in sectors_assets
    assert ("major_banks", "residential_mortgage") in sectors_assets
    assert any("Tab 4a" in r.message for r in caplog.records)


def test_adapter_empty_workbook_returns_empty_valid_df(
    tmp_path: Path, caplog,
) -> None:
    import logging
    caplog.set_level(logging.WARNING)
    path = tmp_path / "empty.xlsx"
    wb = openpyxl.Workbook(); wb.active.title = "Cover"
    wb.save(path)
    adapter = ApraQpexAdapter()
    df = adapter.normalise(path)
    assert df.empty
    for col in adapter.canonical_columns:
        assert col in df.columns


def test_adapter_label_lookup_tolerates_row_offset(live_qpex_workbook: Path) -> None:
    """Tab 1a uses row 26/43; Tab 2a uses 25/42. Both must resolve."""
    df = ApraQpexAdapter().normalise(live_qpex_workbook)
    tab1a = df[df["_source_sheet"] == "Tab 1a"]
    tab2a = df[df["_source_sheet"] == "Tab 2a"]
    assert tab1a["_denominator_row"].unique().tolist() == [26]
    assert tab1a["_numerator_row"].unique().tolist() == [43]
    assert tab2a["_denominator_row"].unique().tolist() == [25]
    assert tab2a["_numerator_row"].unique().tolist() == [42]


def test_validate_output_rejects_missing_column() -> None:
    adapter = ApraQpexAdapter()
    bad = pd.DataFrame({"institution_sector": ["banks"]})
    with pytest.raises(ValueError, match="missing required columns"):
        adapter.validate_output(bad)


def test_adapter_excludes_of_which_subrows(live_qpex_workbook: Path) -> None:
    """The commercial sheet's 'of which' sub-row must not be picked as total."""
    df = ApraQpexAdapter().normalise(live_qpex_workbook)
    # For Tab 1a commercial, denominator must be row 26 (not 27, the "of which").
    tab1a_com = df[df["_source_sheet"] == "Tab 1a"]
    assert not tab1a_com.empty
    assert (tab1a_com["_denominator_row"] == 26).all()
