"""Tests for the CBA quarterly APS 330 XLSX adapter (Option A)."""
from __future__ import annotations

from datetime import date
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from ingestion.adapters.cba_pillar3_quarterly_adapter import (
    CbaPillar3QuarterlyAdapter,
)


def _build_quarterly_workbook(path: Path, *, include_portfolio_sheet: bool = True) -> None:
    """Synthesise a minimal quarterly-shape workbook with CRB(f) + EAD & CRWA."""
    wb = openpyxl.Workbook()
    wb.active.title = "Cover"
    wb.create_sheet("Contents")

    # EAD & CRWA — column 1 carries the EAD $M value.
    ead = wb.create_sheet("EAD & CRWA")
    ead.cell(row=8, column=1, value="Subject to AIRB approach")
    ead.cell(row=9, column=1, value="  Corporate (incl. SME corporate) ")
    ead.cell(row=9, column=2, value=188057)
    ead.cell(row=10, column=1, value="  SME retail")
    ead.cell(row=10, column=2, value=19908)
    ead.cell(row=11, column=1, value="  Residential mortgage ")
    ead.cell(row=11, column=2, value=704075)
    ead.cell(row=12, column=1, value="  Qualifying revolving retail")
    ead.cell(row=12, column=2, value=22826)
    ead.cell(row=20, column=1, value="Specialised lending")
    ead.cell(row=20, column=2, value=5654)

    # CRB(f) — column 1 carries the non-performing $M value.
    crb = wb.create_sheet("CRB(f)")
    if include_portfolio_sheet:
        crb.cell(row=49, column=1, value="Portfolio")
        crb.cell(row=50, column=1, value="Corporate (incl. Large and SME corporate)")
        crb.cell(row=50, column=2, value=2001)
        crb.cell(row=53, column=1, value="SME Retail")
        crb.cell(row=53, column=2, value=371)
        crb.cell(row=54, column=1, value="Residential Mortgage")
        crb.cell(row=54, column=2, value=6353)
        crb.cell(row=55, column=1, value="Qualifying Revolving Retail")
        crb.cell(row=55, column=2, value=79)
        crb.cell(row=57, column=1, value="Specialised Lending")
        crb.cell(row=57, column=2, value=158)

    wb.save(path)


@pytest.fixture()
def quarterly_workbook(tmp_path: Path) -> Path:
    p = tmp_path / "cba_quarterly.xlsx"
    _build_quarterly_workbook(p)
    return p


def test_adapter_finds_portfolios_and_computes_ratios(quarterly_workbook: Path) -> None:
    df = CbaPillar3QuarterlyAdapter().normalise(quarterly_workbook, reporting_date=date(2025, 9, 30))
    by_class = {r["asset_class"]: r for _, r in df.iterrows()}
    assert set(by_class) >= {
        "corporate_aggregate", "retail_sme", "residential_mortgage",
        "retail_qrr", "specialised_lending",
    }


def test_adapter_arithmetic_is_correct(quarterly_workbook: Path) -> None:
    df = CbaPillar3QuarterlyAdapter().normalise(quarterly_workbook, reporting_date=date(2025, 9, 30))
    row = df[df["asset_class"] == "residential_mortgage"].iloc[0]
    assert row["value"] == pytest.approx(6353.0 / 704075.0, rel=1e-9)
    assert row["numerator_value"] == 6353.0
    assert row["denominator_value"] == 704075.0


def test_adapter_preserves_sheet_provenance(quarterly_workbook: Path) -> None:
    df = CbaPillar3QuarterlyAdapter().normalise(quarterly_workbook, reporting_date=date(2025, 9, 30))
    assert (df["numerator_sheet"] == "CRB(f)").all()
    assert (df["denominator_sheet"] == "EAD & CRWA").all()


def test_adapter_empty_when_sheets_missing(tmp_path: Path, caplog) -> None:
    import logging
    caplog.set_level(logging.WARNING)
    p = tmp_path / "empty.xlsx"
    wb = openpyxl.Workbook(); wb.active.title = "Cover"; wb.save(p)
    df = CbaPillar3QuarterlyAdapter().normalise(p, reporting_date=date(2025, 9, 30))
    assert df.empty
    for col in CbaPillar3QuarterlyAdapter()._CANONICAL_COLUMNS:
        assert col in df.columns
    assert any("missing required sheets" in r.message for r in caplog.records)


def test_adapter_filters_implausible_ratio(tmp_path: Path, caplog) -> None:
    """An NPE-to-EAD ratio > 10% must be dropped with a warning."""
    import logging
    caplog.set_level(logging.WARNING)
    p = tmp_path / "implausible.xlsx"
    wb = openpyxl.Workbook(); wb.active.title = "Cover"
    ead = wb.create_sheet("EAD & CRWA")
    ead.cell(row=9, column=1, value="Residential mortgage")
    ead.cell(row=9, column=2, value=1000)
    crb = wb.create_sheet("CRB(f)")
    crb.cell(row=50, column=1, value="Residential Mortgage")
    crb.cell(row=50, column=2, value=500)  # 50% NPL — implausible
    wb.save(p)

    df = CbaPillar3QuarterlyAdapter().normalise(p, reporting_date=date(2025, 9, 30))
    assert df.empty
    assert any("implausible" in r.message.lower() for r in caplog.records)


def test_adapter_handles_missing_portfolio(tmp_path: Path) -> None:
    """If a portfolio is listed in EAD but not in CRB(f), it is silently skipped."""
    p = tmp_path / "partial.xlsx"
    wb = openpyxl.Workbook(); wb.active.title = "Cover"
    ead = wb.create_sheet("EAD & CRWA")
    ead.cell(row=9, column=1, value="Residential mortgage")
    ead.cell(row=9, column=2, value=100000)
    ead.cell(row=10, column=1, value="Sovereign")
    ead.cell(row=10, column=2, value=140000)
    crb = wb.create_sheet("CRB(f)")
    crb.cell(row=50, column=1, value="Residential Mortgage")
    crb.cell(row=50, column=2, value=1000)
    # No Sovereign row in CRB(f).
    wb.save(p)

    df = CbaPillar3QuarterlyAdapter().normalise(p, reporting_date=date(2025, 9, 30))
    assert set(df["asset_class"]) == {"residential_mortgage"}


def test_validate_output_rejects_missing_column() -> None:
    adapter = CbaPillar3QuarterlyAdapter()
    with pytest.raises(ValueError, match="missing required columns"):
        adapter.validate_output(pd.DataFrame({"asset_class": ["x"]}))
