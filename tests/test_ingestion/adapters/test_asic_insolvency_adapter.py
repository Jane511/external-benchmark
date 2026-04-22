"""Tests for the ASIC Insolvency Series 1+2 workbook adapter."""
from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from ingestion.adapters.asic_insolvency_adapter import AsicInsolvencyAdapter


# Column layout per ASIC's Series 1 & 2 workbook. Header sits on row 7
# (0-indexed 6) — the adapter reads the sheet with ``header=6``.
_HEADERS = [
    "Data to",
    "ACN No ",
    "Organisation name ",
    "Appointee (person or company)",
    "Appointment type",
    "Effective date",
    "Period\n(Year month)",
    "Period (financial year)",
    "Industry type (division)",
    "Industry type (subdivision)",
    "Industry type (group)",
    "State of incorporation (state or territory)",
    "Principal place of business (state or territory)",
    "Principal place of business (area)",
    "Principal place of business (postcode)",
    "Series 1 \n(companies entering)",
    "Series 2 \n(all appointments)",
]


def _append_data_row(
    ws, *, appointment: str, industry: str, fy: str, series1: str,
) -> None:
    ws.append([
        None,                                       # Data to
        12345678,                                   # ACN
        "TEST PTY LTD",                             # Organisation
        "APPOINTEE X",                              # Appointee
        appointment,                                # Appointment type
        datetime(2023, 10, 1),                      # Effective date
        "2023 10",                                  # Period (Year month)
        fy,                                         # Period (financial year)
        industry,                                   # Industry division
        "", "", "VIC",                              # subdivision/group/state
        "", "", "",                                 # principal place of business
        series1,                                    # Series 1 flag
        "Yes",                                      # Series 2 flag
    ])


def _build_asic_workbook(path: Path) -> None:
    wb = openpyxl.Workbook()
    wb.active.title = "Contents"
    ws = wb.create_sheet("Data set")
    # Rows 1-6 are header scaffolding — the adapter reads with header=6.
    for _ in range(6):
        ws.append([None] * len(_HEADERS))
    ws.append(_HEADERS)

    # FY2024 Construction: 3 Series-1 "Yes" + 1 subsequent "No"
    for app in ("Voluntary administration", "Court liquidation",
                "Creditors' voluntary liquidation"):
        _append_data_row(ws, appointment=app, industry="Construction",
                         fy="2023-2024", series1="Yes")
    _append_data_row(ws, appointment="Receiver appointed", industry="Construction",
                     fy="2023-2024", series1="No")

    # FY2024 Retail Trade: 2 Series-1
    for _ in range(2):
        _append_data_row(ws, appointment="Creditors' voluntary liquidation",
                         industry="Retail Trade", fy="2023-2024", series1="Yes")

    # FY2024 "Unknown" industry — must be filtered with warning.
    _append_data_row(ws, appointment="Voluntary administration",
                     industry="Unknown", fy="2023-2024", series1="Yes")

    # FY2025 Construction: 1 Series-1
    _append_data_row(ws, appointment="Voluntary administration",
                     industry="Construction", fy="2024-2025", series1="Yes")

    wb.save(path)


@pytest.fixture()
def asic_workbook(tmp_path: Path) -> Path:
    p = tmp_path / "Series-1-2.xlsx"
    _build_asic_workbook(p)
    return p


def test_adapter_filters_to_series_1_yes_only(asic_workbook: Path) -> None:
    df = AsicInsolvencyAdapter().normalise(asic_workbook)
    # Expected groups (Series 1 Yes + known division):
    #   FY2024 Construction = 3, FY2024 Retail Trade = 2, FY2025 Construction = 1
    counts = {
        (row["industry"], row["fiscal_year"]): row["insolvency_count"]
        for _, row in df.iterrows()
    }
    assert counts[("Construction", "FY2024")] == 3
    assert counts[("Retail Trade", "FY2024")] == 2
    assert counts[("Construction", "FY2025")] == 1
    # Total rows = 3 groups.
    assert len(df) == 3


def test_adapter_drops_unknown_industry_with_warning(
    asic_workbook: Path, caplog,
) -> None:
    import logging
    caplog.set_level(logging.WARNING)
    df = AsicInsolvencyAdapter().normalise(asic_workbook)
    assert "Unknown" not in set(df["industry"])
    assert any("unmapped industry" in r.message.lower() for r in caplog.records)


def test_adapter_converts_fy_label_to_end_date(asic_workbook: Path) -> None:
    df = AsicInsolvencyAdapter().normalise(asic_workbook)
    fy2024 = df[df["fiscal_year"] == "FY2024"].iloc[0]
    assert fy2024["as_of_date"] == date(2024, 6, 30)


def test_adapter_records_filter_in_provenance(asic_workbook: Path) -> None:
    df = AsicInsolvencyAdapter().normalise(asic_workbook)
    assert (df["filter_applied"] == "Series 1 (companies entering)").all()
    assert (df["source_sheet"] == "Data set").all()


def test_adapter_handles_missing_columns(tmp_path: Path, caplog) -> None:
    import logging
    caplog.set_level(logging.WARNING)
    p = tmp_path / "bad.xlsx"
    wb = openpyxl.Workbook(); wb.active.title = "Data set"
    # Header row is there but missing Series 1 column.
    wb["Data set"].append(["Industry type (division)", "Period (financial year)"])
    wb.save(p)
    df = AsicInsolvencyAdapter().normalise(p)
    assert df.empty


def test_adapter_produces_canonical_columns(asic_workbook: Path) -> None:
    adapter = AsicInsolvencyAdapter()
    df = adapter.normalise(asic_workbook)
    for col in adapter.canonical_columns:
        assert col in df.columns


def test_validate_output_rejects_missing_column() -> None:
    adapter = AsicInsolvencyAdapter()
    with pytest.raises(ValueError, match="missing required columns"):
        adapter.validate_output(pd.DataFrame({"industry": ["x"]}))


def test_adapter_logs_filter_statistics(asic_workbook: Path, caplog) -> None:
    import logging
    caplog.set_level(logging.INFO)
    AsicInsolvencyAdapter().normalise(asic_workbook)
    assert any("Series 1 filter retained" in r.message for r in caplog.records)
