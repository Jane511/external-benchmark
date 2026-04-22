"""Tests for the live-APRA Performance adapter.

Builds a small synthetic workbook that mirrors the real APRA layout
(wide-format ``Tab Xd`` sheets with dates as column headers and APS 220
ratio rows at known positions). No real APRA file is committed.
"""
from __future__ import annotations

from datetime import date
from pathlib import Path

import openpyxl
import pytest

from ingestion.adapters.apra_performance_adapter import (
    SYNTHETIC_ASSET_CLASS,
    ApraPerformanceAdapter,
)


# ---------------------------------------------------------------------------
# Synthetic workbook builder
# ---------------------------------------------------------------------------

_DATE_COLUMNS = [
    "Jun 2021", "Sep 2021", "Dec 2021",            # pre-APS-220 (should skip)
    "Mar 2022", "Jun 2022", "Sep 2022", "Dec 2022",
    "Mar 2023", "Jun 2023", "Sep 2023", "Dec 2023",
    "Mar 2024", "Jun 2024", "Sep 2024", "Dec 2024",
]


def _build_sector_sheet(
    ws,
    *,
    title: str,
    npl_values: list[float | None],
    dpd_values: list[float | None],
    date_headers: list[str] = _DATE_COLUMNS,
) -> None:
    """Populate a worksheet with the APRA Tab Xd shape."""
    # Row 0 (index 0 = Excel row 1)
    ws.cell(row=1, column=1, value=title)
    # Row 1 — unit line (unused)
    ws.cell(row=2, column=1, value="($ million, consolidated group)")
    # Row 2 — banner
    ws.cell(row=3, column=2, value="Quarter end")
    # Row 3 — date headers in columns 2..N (Excel cols 2..N+1)
    for i, hdr in enumerate(date_headers):
        ws.cell(row=4, column=2 + i, value=hdr)
    # Row 53 (index 53 = Excel row 54) — NPL ratio
    ws.cell(row=54, column=1, value="Non-performing to loans and advances")
    for i, v in enumerate(npl_values):
        if v is not None:
            ws.cell(row=54, column=2 + i, value=v)
    # Row 54 — 90+DPD ratio
    ws.cell(row=55, column=1,
            value="  of which >= 90 days past due to loans and advances")
    for i, v in enumerate(dpd_values):
        if v is not None:
            ws.cell(row=55, column=2 + i, value=v)


@pytest.fixture()
def live_apra_workbook(tmp_path: Path) -> Path:
    """Synthetic workbook with Tab 1d, 2d, 4d populated + pre-2022 blanks."""
    path = tmp_path / "live_apra_performance.xlsx"
    wb = openpyxl.Workbook()

    # Default sheet — make it a "Cover" front-matter sheet
    wb.active.title = "Cover"

    # Tab 1d (All ADIs)
    ws = wb.create_sheet("Tab 1d")
    npl = [None, None, None] + [0.008, 0.009, 0.009, 0.010,
                                0.010, 0.011, 0.011, 0.011,
                                0.012, 0.012, 0.012, 0.011]
    dpd = [None, None, None] + [0.005, 0.005, 0.005, 0.006,
                                0.006, 0.007, 0.007, 0.006,
                                0.007, 0.007, 0.006, 0.006]
    _build_sector_sheet(ws, title="Table 1d  All ADIs' asset quality",
                        npl_values=npl, dpd_values=dpd)

    # Tab 2d (Banks)
    ws = wb.create_sheet("Tab 2d")
    _build_sector_sheet(ws, title="Table 2d  Banks' asset quality",
                        npl_values=npl, dpd_values=dpd)

    # Tab 4d (Major banks)
    ws = wb.create_sheet("Tab 4d")
    _build_sector_sheet(ws, title="Table 4d  Major banks' asset quality",
                        npl_values=npl, dpd_values=dpd)

    wb.save(path)
    return path


@pytest.fixture()
def live_apra_with_implausible(tmp_path: Path) -> Path:
    """Like ``live_apra_workbook`` but one cell holds an implausible 0.5 (50%)."""
    path = tmp_path / "live_apra_bad.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "Cover"

    ws = wb.create_sheet("Tab 2d")
    npl = [None, None, None, 0.008, 0.009, 0.009, 0.5,  # <- 50% NPL, implausible
           0.010, 0.010, 0.011, 0.011, 0.011, 0.012, 0.012, 0.012]
    dpd = [None, None, None, 0.005, 0.005, 0.005, 0.006,
           0.006, 0.007, 0.007, 0.006, 0.007, 0.007, 0.006, 0.006]
    _build_sector_sheet(ws, title="Table 2d  Banks' asset quality",
                        npl_values=npl, dpd_values=dpd)
    wb.save(path)
    return path


@pytest.fixture()
def live_apra_empty_workbook(tmp_path: Path) -> Path:
    """Workbook that has no Tab Xd sheets at all."""
    path = tmp_path / "live_apra_empty.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "Cover"
    wb.create_sheet("Notes")
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


def test_adapter_finds_sector_sheets(live_apra_workbook: Path) -> None:
    df = ApraPerformanceAdapter().normalise(live_apra_workbook)
    assert set(df["institution_sector"].unique()) == {
        "all_adis", "banks", "major_banks",
    }


def test_adapter_filters_pre_2022_blanks(live_apra_workbook: Path) -> None:
    df = ApraPerformanceAdapter().normalise(live_apra_workbook)
    # 12 APS-220-populated quarters in the synthetic workbook (Mar 2022..Dec 2024)
    # × 2 metrics × 3 sectors = 72 rows; no 2021 quarters should appear.
    assert len(df) == 72
    assert all(r >= date(2022, 3, 31) for r in df["as_of_date"])


def test_adapter_extracts_metric_rows(live_apra_workbook: Path) -> None:
    df = ApraPerformanceAdapter().normalise(live_apra_workbook)
    assert set(df["metric_name"].unique()) == {"npl_ratio", "ninety_dpd_rate"}
    # All NPL values in the synthetic workbook sit in [0.008, 0.012]
    npl_vals = df.loc[df["metric_name"] == "npl_ratio", "value"]
    assert npl_vals.min() >= 0.008
    assert npl_vals.max() <= 0.012


def test_adapter_produces_canonical_columns(live_apra_workbook: Path) -> None:
    adapter = ApraPerformanceAdapter()
    df = adapter.normalise(live_apra_workbook)
    for col in adapter.canonical_columns:
        assert col in df.columns


def test_adapter_snaps_dates_to_quarter_end(live_apra_workbook: Path) -> None:
    df = ApraPerformanceAdapter().normalise(live_apra_workbook)
    # "Mar 2022" should become 2022-03-31, "Jun 2022" → 2022-06-30, etc.
    expected_days = {3: 31, 6: 30, 9: 30, 12: 31}
    for d in df["as_of_date"]:
        assert d.day == expected_days[d.month], d


def test_adapter_period_slugs_line_up_with_dates(live_apra_workbook: Path) -> None:
    df = ApraPerformanceAdapter().normalise(live_apra_workbook)
    for _, row in df.head(10).iterrows():
        d = row["as_of_date"]
        expected = f"{d.year}Q{(d.month - 1) // 3 + 1}"
        assert row["period"] == expected


def test_adapter_drops_implausible_values(
    live_apra_with_implausible: Path, caplog,
) -> None:
    import logging
    caplog.set_level(logging.WARNING)
    df = ApraPerformanceAdapter().normalise(live_apra_with_implausible)
    # The 0.5 NPL row must be filtered; all remaining NPLs in [0, 0.1].
    assert (df.loc[df["metric_name"] == "npl_ratio", "value"] <= 0.10).all()
    assert any("implausible" in rec.message.lower() for rec in caplog.records)


def test_adapter_returns_empty_canonical_frame_when_no_sheets(
    live_apra_empty_workbook: Path, caplog,
) -> None:
    import logging
    caplog.set_level(logging.WARNING)
    adapter = ApraPerformanceAdapter()
    df = adapter.normalise(live_apra_empty_workbook)
    assert df.empty
    # Canonical columns still present so downstream code can iterate safely.
    for col in adapter.canonical_columns:
        assert col in df.columns
    # Every configured sheet should have produced a warning.
    missing_messages = [
        r for r in caplog.records if "not in workbook" in r.message
    ]
    assert len(missing_messages) >= len(ApraPerformanceAdapter.SECTOR_SHEETS)


def test_adapter_asset_class_is_synthetic_label(
    live_apra_workbook: Path,
) -> None:
    df = ApraPerformanceAdapter().normalise(live_apra_workbook)
    assert (df["asset_class"] == SYNTHETIC_ASSET_CLASS).all()


def test_validate_output_rejects_missing_column() -> None:
    import pandas as pd
    adapter = ApraPerformanceAdapter()
    bad = pd.DataFrame({"institution_sector": ["banks"]})
    with pytest.raises(ValueError, match="missing required columns"):
        adapter.validate_output(bad)
