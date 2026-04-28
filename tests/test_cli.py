"""Smoke tests for cli.py — verify each subcommand runs over a seeded DB."""
from __future__ import annotations

import json
import sys
from datetime import date
from pathlib import Path

import pytest
from click.testing import CliRunner

# cli.py lives at the project root, one level up from tests/
PROJECT_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(PROJECT_ROOT))
from cli import cli  # noqa: E402


@pytest.fixture()
def db_path(tmp_path: Path) -> str:
    return str(tmp_path / "bench.db")


def _invoke(runner: CliRunner, db: str, *args: str):
    return runner.invoke(cli, ["--db", db, *args])


def test_seed_command_creates_entries(db_path) -> None:
    runner = CliRunner()
    result = _invoke(runner, db_path, "seed")
    assert result.exit_code == 0
    assert "Loaded" in result.output


def test_list_after_seed(db_path) -> None:
    runner = CliRunner()
    _invoke(runner, db_path, "seed")
    result = _invoke(runner, db_path, "list")
    assert result.exit_code == 0
    assert "CBA_PILLAR3_RES_2024H2" in result.output


def test_history_shows_version_chain(db_path) -> None:
    runner = CliRunner()
    _invoke(runner, db_path, "seed")
    result = _invoke(runner, db_path, "history", "CBA_PILLAR3_RES_2024H2")
    assert result.exit_code == 0
    assert "v1" in result.output


def test_report_stale(db_path) -> None:
    runner = CliRunner()
    _invoke(runner, db_path, "seed")
    result = _invoke(runner, db_path, "report", "stale")
    assert result.exit_code == 0
    assert "stale_benchmarks" in result.output


def test_report_coverage_requires_segment(db_path) -> None:
    runner = CliRunner()
    _invoke(runner, db_path, "seed")
    result = _invoke(runner, db_path, "report", "coverage")
    assert result.exit_code != 0
    assert "requires --segment" in result.output


def test_report_coverage_with_segment(db_path) -> None:
    runner = CliRunner()
    _invoke(runner, db_path, "seed")
    result = _invoke(
        runner, db_path, "report", "coverage", "--segment", "residential_mortgage",
    )
    assert result.exit_code == 0
    assert "coverage" in result.output


def test_feed_command_now_deprecated(db_path) -> None:
    """The `feed` command was replaced by `observations` in Brief 1."""
    runner = CliRunner()
    _invoke(runner, db_path, "seed")
    result = _invoke(
        runner, db_path, "feed", "central_tendency",
        "--segment", "residential_mortgage",
    )
    assert result.exit_code != 0
    assert "deprecated" in (result.output + (result.stderr or "")).lower()


def test_observations_command_runs(db_path) -> None:
    runner = CliRunner()
    _invoke(runner, db_path, "seed")
    result = _invoke(
        runner, db_path, "observations",
        "--segment", "residential_mortgage",
        "--format", "table",
    )
    # Empty seed (legacy BenchmarkEntry rows) is fine — exit must still be 0.
    assert result.exit_code == 0


def test_export_json_stdout(db_path) -> None:
    runner = CliRunner()
    _invoke(runner, db_path, "seed")
    result = _invoke(runner, db_path, "export", "--format", "json")
    assert result.exit_code == 0
    records = json.loads(result.output)
    assert len(records) >= 30  # seed entries (LGD components extracted to future_lgd/)


def test_export_csv_to_file(db_path, tmp_path) -> None:
    runner = CliRunner()
    _invoke(runner, db_path, "seed")
    out_csv = tmp_path / "export.csv"
    result = _invoke(
        runner, db_path, "export", "--format", "csv", "--output", str(out_csv),
    )
    assert result.exit_code == 0
    assert out_csv.exists()
    assert "source_id,version," in out_csv.read_text()


# ---------------------------------------------------------------------------
# Ingest subcommands (Phase 1)
# ---------------------------------------------------------------------------

def _make_apra_fixture(path: Path) -> None:
    """Write a 3-row APRA-shaped XLSX at `path` for CLI tests."""
    import openpyxl
    from datetime import date as _date

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Asset Quality"
    ws.append(["Period", "Category", "90DPD_Rate", "NPL_Rate"])
    ws.append([_date(2025, 9, 30), "Residential", 0.012, 0.008])
    ws.append([_date(2025, 9, 30), "Commercial", 0.018, 0.012])
    ws.append([_date(2025, 9, 30), "Corporate", 0.025, 0.015])
    wb.save(path)


def test_ingest_apra_dry_run_prints_actions_without_writing(db_path, tmp_path) -> None:
    runner = CliRunner()
    xlsx = tmp_path / "apra.xlsx"
    _make_apra_fixture(xlsx)

    result = _invoke(
        runner, db_path, "ingest", "apra",
        "--source-path", str(xlsx),
        "--dry-run",
    )
    assert result.exit_code == 0
    assert "dry_run=True" in result.output
    assert "add=6" in result.output         # 3 categories x 2 metrics

    # Verify nothing was written.
    listed = _invoke(runner, db_path, "list")
    assert "0 entries" in listed.output or listed.output.strip().endswith("0 entries.")


def test_ingest_apra_writes_entries_on_real_run(db_path, tmp_path) -> None:
    runner = CliRunner()
    xlsx = tmp_path / "apra.xlsx"
    _make_apra_fixture(xlsx)

    result = _invoke(
        runner, db_path, "ingest", "apra",
        "--source-path", str(xlsx),
    )
    assert result.exit_code == 0
    assert "add=6" in result.output

    listed = _invoke(runner, db_path, "list")
    assert listed.exit_code == 0
    assert "APRA_RESIDENTIAL_MORTGAGE_90DPD_RATE" in listed.output


def test_ingest_apra_twice_is_idempotent(db_path, tmp_path) -> None:
    runner = CliRunner()
    xlsx = tmp_path / "apra.xlsx"
    _make_apra_fixture(xlsx)

    _invoke(runner, db_path, "ingest", "apra", "--source-path", str(xlsx))
    second = _invoke(runner, db_path, "ingest", "apra", "--source-path", str(xlsx))
    assert second.exit_code == 0
    assert "skip_unchanged=6" in second.output


def test_ingest_status_shows_apra_after_ingest(db_path, tmp_path) -> None:
    runner = CliRunner()
    xlsx = tmp_path / "apra.xlsx"
    _make_apra_fixture(xlsx)
    _invoke(runner, db_path, "ingest", "apra", "--source-path", str(xlsx))

    status = _invoke(runner, db_path, "ingest", "status")
    assert status.exit_code == 0
    assert "apra_adi" in status.output


# ---------------------------------------------------------------------------
# Ingest pillar3 subcommands (Phase 2)
# ---------------------------------------------------------------------------

def _make_cba_fixture(path: Path) -> None:
    """Write a CBA Pillar 3 Excel companion fixture matching sources.yaml."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "IRB Credit Risk"
    ws.append(["Portfolio", "Exposure_EAD_Mn", "PD", "LGD"])
    ws.append(["Residential Mortgage", 500000, 0.0072, 0.22])
    ws.append(["CRE Investment",       80000,  0.0250, 0.35])
    ws.append(["Corporate SME",        120000, 0.0280, 0.45])

    sl = wb.create_sheet("Specialised Lending")
    sl.append(["Grade", "PD"])
    sl.append(["Strong",       0.004])
    sl.append(["Good",         0.008])
    sl.append(["Satisfactory", 0.028])
    sl.append(["Weak",         0.080])
    wb.save(path)


def test_ingest_pillar3_cba_dry_run(db_path, tmp_path) -> None:
    runner = CliRunner()
    xlsx = tmp_path / "cba.xlsx"
    _make_cba_fixture(xlsx)

    result = _invoke(
        runner, db_path, "ingest", "pillar3", "cba",
        "--source-path", str(xlsx),
        "--reporting-date", "2025-06-30",
        "--dry-run",
    )
    assert result.exit_code == 0
    assert "dry_run=True" in result.output
    assert "add=10" in result.output                         # 6 IRB + 4 slotting
    assert "CBA_RESIDENTIAL_MORTGAGE_PD_FY2025" in result.output

    listed = _invoke(runner, db_path, "list")
    assert "0 entries" in listed.output or listed.output.strip().endswith("0 entries.")


def test_ingest_pillar3_cba_writes_entries(db_path, tmp_path) -> None:
    runner = CliRunner()
    xlsx = tmp_path / "cba.xlsx"
    _make_cba_fixture(xlsx)

    result = _invoke(
        runner, db_path, "ingest", "pillar3", "cba",
        "--source-path", str(xlsx),
        "--reporting-date", "2025-06-30",
    )
    assert result.exit_code == 0
    assert "add=10" in result.output

    listed = _invoke(runner, db_path, "list")
    assert "CBA_RESIDENTIAL_MORTGAGE_PD_FY2025" in listed.output
    assert "CBA_DEVELOPMENT_STRONG_PD_FY2025" in listed.output


def test_ingest_pillar3_without_bank_runs_all_four_banks(db_path, tmp_path) -> None:
    """With a --cba-path provided, CBA runs; NAB/WBC/ANZ fall through to the
    FileDownloader (which we mock so no network call)."""
    import openpyxl
    from unittest.mock import patch

    runner = CliRunner()
    xlsx = tmp_path / "cba.xlsx"
    _make_cba_fixture(xlsx)

    # Minimal PDF-ish stub content; pdfplumber opens it in scrape(), so we also
    # need to avoid actual PDF parsing for the PDF banks. Easiest path: mock
    # the downloader to write a dummy file, then assert the scrape raises
    # (empty PDF) and the orchestrator reports the error cleanly without
    # taking down the run. The group command reports each bank's outcome.
    def fake_urlretrieve(_url, dest):
        Path(dest).write_bytes(b"%PDF-1.4\n%%EOF\n")   # minimally valid PDF header
        return dest, None

    with patch("ingestion.downloader.urlretrieve", side_effect=fake_urlretrieve):
        result = _invoke(
            runner, db_path, "ingest", "pillar3",
            "--cba-path", str(xlsx),
            "--reporting-date", "2025-06-30",
            "--dry-run",
        )

    assert result.exit_code == 0
    # Headers for all four banks — no "not yet implemented" anywhere.
    for bank in ("--- CBA ---", "--- NAB ---", "--- WBC ---", "--- ANZ ---"):
        assert bank in result.output
    assert "not yet implemented" not in result.output
    # CBA should succeed; NAB/WBC/ANZ will report parse errors or zero results
    # from the trivial dummy PDF, which is expected without a real PDF.
    assert "add=10" in result.output   # the CBA slice


def test_ingest_pillar3_without_bank_skips_cba_if_no_path(db_path) -> None:
    """Group mode with no --cba-path still attempts all four; CBA path tries
    the FileDownloader, which in test would require a mock. Without a mock we
    expect the orchestrator to report errors for all four and exit cleanly."""
    runner = CliRunner()
    result = _invoke(
        runner, db_path, "ingest", "pillar3",
        "--dry-run",
    )
    assert result.exit_code == 0
    # All four banks get a header, even if they error later
    for bank in ("--- CBA ---", "--- NAB ---", "--- WBC ---", "--- ANZ ---"):
        assert bank in result.output


# ---------------------------------------------------------------------------
# NAB / WBC / ANZ subcommands (Phase 2 completion)
# ---------------------------------------------------------------------------

def test_ingest_pillar3_nab_reports_error_without_real_pdf(db_path, tmp_path) -> None:
    """NAB subcommand now exists (no longer NotImplementedError). A trivial
    dummy PDF will fail pdfplumber parsing — orchestrator catches + reports."""
    runner = CliRunner()
    fake_pdf = tmp_path / "nab.pdf"
    fake_pdf.write_bytes(b"%PDF-1.4\n%%EOF\n")

    result = _invoke(
        runner, db_path, "ingest", "pillar3", "nab",
        "--source-path", str(fake_pdf),
        "--reporting-date", "2025-06-30",
        "--dry-run",
    )
    # Exit code may be non-zero (parse error), but we must NOT see NotImplementedError
    assert "NotImplementedError" not in result.output
    assert "not yet implemented" not in result.output


def test_ingest_pillar3_anz_subcommand_is_registered(db_path) -> None:
    """Smoke: `ingest pillar3 anz --help` should succeed, proving the command exists."""
    runner = CliRunner()
    result = runner.invoke(cli, ["ingest", "pillar3", "anz", "--help"])
    assert result.exit_code == 0
    assert "ANZ" in result.output or "anz" in result.output.lower()


# ---------------------------------------------------------------------------
# Ingest ICC Trade Register (Phase 3)
# ---------------------------------------------------------------------------

def test_ingest_icc_subcommand_reports_missing_cache_dir(db_path) -> None:
    """Without a cached PDF, `ingest icc` should surface the manual-download hint."""
    runner = CliRunner()
    # Use isolated_filesystem so data/raw/icc/ resolves to a tmp dir
    with runner.isolated_filesystem() as fs:
        result = runner.invoke(
            cli,
            ["--db", str(Path(fs) / "b.db"),
             "ingest", "icc", "--dry-run"],
        )
    # Exit code non-zero (FileNotFoundError surfaces as an orchestrator error)
    # but more importantly, we should see a clear manual-download hint.
    assert "icc" in result.output.lower()
    # The error should mention manual-download or the iccwbo URL.
    assert (
        "manual-download" in result.output.lower()
        or "iccwbo" in result.output.lower()
        or "not found" in result.output.lower()
    )


def test_ingest_icc_help_shows_force_refresh_and_report_year(db_path) -> None:
    """Verify the command is registered with the expected flags."""
    runner = CliRunner()
    result = runner.invoke(cli, ["ingest", "icc", "--help"])
    assert result.exit_code == 0
    assert "--report-year" in result.output
    assert "--force-refresh" in result.output
    assert "ICC" in result.output




# ---------------------------------------------------------------------------
# Report 1 committee report CLI (Part 3)
# ---------------------------------------------------------------------------

def test_report_benchmark_markdown_writes_raw_only_file(db_path, tmp_path) -> None:
    """Markdown emits a single raw-only file with the new section structure."""
    runner = CliRunner()
    _invoke(runner, db_path, "seed")
    out = tmp_path / "report.md"
    result = _invoke(
        runner, db_path, "report", "benchmark",
        "--format", "markdown",
        "--output", str(out),
        "--period-label", "Q3 2025",
    )
    assert result.exit_code == 0
    text = out.read_text(encoding="utf-8")
    assert "## 1. Executive Summary" in text
    assert "## 2. Per-source raw observations by segment" in text
    assert "raw, source-attributable observations only" in text


def test_report_benchmark_html_writes_self_contained_file(db_path, tmp_path) -> None:
    runner = CliRunner()
    _invoke(runner, db_path, "seed")
    out = tmp_path / "report.html"
    result = _invoke(
        runner, db_path, "report", "benchmark",
        "--format", "html",
        "--output", str(out),
        "--period-label", "Q3 2025",
    )
    assert result.exit_code == 0
    text = out.read_text(encoding="utf-8")
    assert "<style>" in text
    assert "raw, source-attributable observations only" in text


def test_report_benchmark_institution_type_flag_is_deprecated(db_path, tmp_path) -> None:
    """--institution-type is accepted for backward compatibility but warned."""
    runner = CliRunner()
    _invoke(runner, db_path, "seed")
    out = tmp_path / "report.md"
    result = _invoke(
        runner, db_path, "report", "benchmark",
        "--format", "markdown",
        "--institution-type", "bank",
        "--output", str(out),
        "--period-label", "Q3 2025",
    )
    assert result.exit_code == 0
    assert "deprecated" in (result.output + (result.stderr or "")).lower()


# ---------------------------------------------------------------------------
# Cache group
# ---------------------------------------------------------------------------

def test_cache_status_empty_prints_no_cached_files(tmp_path) -> None:
    runner = CliRunner()
    result = runner.invoke(
        cli, ["cache", "status", "--cache-base", str(tmp_path)],
    )
    assert result.exit_code == 0
    assert "No cached files" in result.output


def test_cache_status_shows_populated_subdir(tmp_path) -> None:
    apra = tmp_path / "apra"
    apra.mkdir()
    (apra / "ADI_Performance_Q3_2025.xlsx").write_bytes(b"x")

    runner = CliRunner()
    result = runner.invoke(
        cli, ["cache", "status", "--cache-base", str(tmp_path)],
    )
    assert result.exit_code == 0
    assert "apra" in result.output
    assert "ADI_Performance_Q3_2025.xlsx" in result.output


def test_cache_clear_specific_source(tmp_path) -> None:
    apra = tmp_path / "apra"
    apra.mkdir()
    (apra / "one.xlsx").write_bytes(b"x")
    (apra / "two.xlsx").write_bytes(b"y")

    runner = CliRunner()
    result = runner.invoke(
        cli, ["cache", "clear", "--source", "apra",
              "--cache-base", str(tmp_path)],
    )
    assert result.exit_code == 0
    assert "Cleared 2" in result.output
    assert list(apra.iterdir()) == []


def test_cache_clear_all_requires_confirmation(tmp_path) -> None:
    apra = tmp_path / "apra"
    apra.mkdir()
    (apra / "one.xlsx").write_bytes(b"x")

    runner = CliRunner()
    # Say "no" at prompt
    result_no = runner.invoke(
        cli, ["cache", "clear", "--cache-base", str(tmp_path)],
        input="n\n",
    )
    assert result_no.exit_code == 0
    assert "Aborted" in result_no.output
    assert list(apra.iterdir())   # not cleared

    # Use --yes to bypass the prompt
    result_yes = runner.invoke(
        cli, ["cache", "clear", "--yes", "--cache-base", str(tmp_path)],
    )
    assert result_yes.exit_code == 0
    assert "Cleared 1" in result_yes.output
    assert list(apra.iterdir()) == []


# ---------------------------------------------------------------------------
# --force-refresh on ingest commands
# ---------------------------------------------------------------------------

def test_ingest_apra_force_refresh_triggers_download(db_path, tmp_path) -> None:
    """With no --source-path and --force-refresh, the scraper hits the downloader."""
    import openpyxl
    from unittest.mock import patch

    runner = CliRunner()
    cache_base = tmp_path / "raw"
    cache_base.mkdir()

    def fake_urlretrieve(url, dest):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Asset Quality"
        ws.append(["Period", "Category", "90DPD_Rate", "NPL_Rate"])
        ws.append([date(2025, 9, 30), "Residential", 0.012, 0.008])
        wb.save(Path(dest))
        return dest, None

    with runner.isolated_filesystem(temp_dir=str(tmp_path)) as fs:
        # Inside the isolated fs, default cache_base=data/raw resolves here.
        with patch("ingestion.downloader.urlretrieve", side_effect=fake_urlretrieve) as mock:
            result = runner.invoke(
                cli,
                ["--db", str(Path(fs) / "b.db"),
                 "ingest", "apra", "--force-refresh", "--dry-run"],
            )
        assert result.exit_code == 0
        assert mock.call_count >= 1


def test_ingest_pillar3_cba_force_refresh_uses_cache_path(db_path, tmp_path) -> None:
    """--force-refresh + no --source-path + --reporting-date routes through FileDownloader."""
    import openpyxl
    from unittest.mock import patch

    runner = CliRunner()

    def fake_urlretrieve(url, dest):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "IRB Credit Risk"
        ws.append(["Portfolio", "Exposure_EAD_Mn", "PD", "LGD"])
        ws.append(["Residential Mortgage", 500000, 0.0072, 0.22])
        wb.save(Path(dest))
        return dest, None

    with runner.isolated_filesystem(temp_dir=str(tmp_path)) as fs:
        with patch("ingestion.downloader.urlretrieve", side_effect=fake_urlretrieve) as mock:
            result = runner.invoke(
                cli,
                ["--db", str(Path(fs) / "b.db"),
                 "ingest", "pillar3", "cba",
                 "--force-refresh",
                 "--reporting-date", "2025-06-30",
                 "--dry-run"],
            )
        assert result.exit_code == 0
        assert mock.call_count == 1
        # Filename should reflect H1 + 2025
        call_args = mock.call_args
        dest_path = Path(call_args.args[1])
        assert "H1" in dest_path.name
        assert "2025" in dest_path.name
