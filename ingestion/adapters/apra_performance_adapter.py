"""APRA Quarterly ADI Performance adapter (Path A — sector-level only).

Live APRA Performance workbooks use a wide-format layout: one sheet per
ADI sector (``Tab 1d`` = All ADIs, ``Tab 2d`` = Banks, ``Tab 4d`` = Major
banks, …), with quarter-end dates as COLUMN headers and metric labels as
row values in column 0. The APS 220-aligned NPL and 90+DPD ratio rows are
only populated from Mar 2022 onward — earlier columns for those rows are
blank.

This adapter extracts those APS 220 ratio rows for a small set of ADI
sectors, reshapes into long format, and returns a canonical DataFrame.
It does NOT attempt asset-class disaggregation — APRA Performance has no
asset-class dimension. Asset-class breakdowns live in QPEX and will be
handled by a separate adapter in a follow-up session (Path B).

Tunables are class attributes so future APRA layout changes are one-line
edits:

- ``SECTOR_SHEETS``       — which ``Tab Xd`` sheets to scan, mapped to a
                            sector code used in ``metric_code`` / audit.
- ``METRIC_ROW_MAP``      — row index → metric key. Row 53 = NPL ratio,
                            row 54 = 90+DPD ratio in the current layout.
- ``DATE_HEADER_ROW``     — row index holding the quarter-end dates. The
                            adapter scans ``_find_date_row`` if the
                            preferred row is empty.
- ``FIRST_APS220_QUARTER``— quarters earlier than this are silently
                            skipped because APS 220 rows are blank.
- ``PLAUSIBILITY``        — per-metric (lo, hi) sanity ranges; values
                            outside are logged and dropped.
"""

from __future__ import annotations

import logging
from datetime import date
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd

from ingestion.adapters.apra_helpers import (
    coerce_quarter_end as _coerce_quarter_end,  # re-export for tests
    date_to_period_slug,
    find_date_row,
    parse_date_row,
)
from ingestion.adapters.base import AbstractAdapter

logger = logging.getLogger(__name__)

# Re-export so external callers / tests keep importing the helper from here.
__all__ = [
    "ApraPerformanceAdapter",
    "SYNTHETIC_ASSET_CLASS",
    "_coerce_quarter_end",
]


SYNTHETIC_ASSET_CLASS = "adi_sector_total"


class ApraPerformanceAdapter(AbstractAdapter):
    """Extract APS 220 sector-level NPL / 90+DPD ratios from live APRA Perf."""

    # ---- adapter contract --------------------------------------------------

    _SOURCE_NAME = "apra_adi_performance"
    _CANONICAL_COLUMNS = [
        "institution_sector",
        "asset_class",
        "period",
        "metric_name",
        "value",
        "as_of_date",
    ]

    # ---- tunables ----------------------------------------------------------

    SECTOR_SHEETS: dict[str, str] = {
        "all_adis": "Tab 1d",
        "banks": "Tab 2d",
        "major_banks": "Tab 4d",
    }

    METRIC_ROW_MAP: dict[int, str] = {
        53: "npl_ratio",
        54: "ninety_dpd_rate",
    }

    DATE_HEADER_ROW: int = 3

    FIRST_APS220_QUARTER: date = date(2022, 3, 31)

    PLAUSIBILITY: dict[str, tuple[float, float]] = {
        "npl_ratio": (0.0, 0.10),
        "ninety_dpd_rate": (0.0, 0.15),
    }

    # ---- AbstractAdapter implementation ------------------------------------

    @property
    def source_name(self) -> str:
        return self._SOURCE_NAME

    @property
    def canonical_columns(self) -> list[str]:
        return list(self._CANONICAL_COLUMNS)

    def normalise(self, file_path: Path) -> pd.DataFrame:
        """Return a long-format DataFrame of sector × metric × quarter rows."""
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        records: list[dict[str, Any]] = []

        missing_sheets: list[str] = []
        for sector_code, sheet_name in self.SECTOR_SHEETS.items():
            if sheet_name not in wb.sheetnames:
                missing_sheets.append(sheet_name)
                logger.warning(
                    "%s: sheet %r not in workbook (available: %s)",
                    type(self).__name__,
                    sheet_name,
                    wb.sheetnames[:8],
                )
                continue

            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))

            try:
                date_row_idx, dates = self._find_date_row(rows)
            except ValueError as exc:
                logger.warning(
                    "%s: %s in sheet %r — skipping",
                    type(self).__name__, exc, sheet_name,
                )
                continue

            for row_idx, metric_name in self.METRIC_ROW_MAP.items():
                if row_idx >= len(rows):
                    logger.warning(
                        "%s: row %d beyond sheet %r (len=%d) — skipping metric %r",
                        type(self).__name__, row_idx, sheet_name, len(rows), metric_name,
                    )
                    continue

                row = rows[row_idx]
                # Values align to the date columns discovered in _find_date_row.
                values = [row[c] if c < len(row) else None for c, _ in dates]

                for (col_idx, as_of), value in zip(dates, values):
                    if value is None or value == "":
                        continue
                    try:
                        value_f = float(value)
                    except (TypeError, ValueError):
                        continue

                    if as_of < self.FIRST_APS220_QUARTER:
                        continue

                    if not self._is_plausible(metric_name, value_f):
                        logger.warning(
                            "%s: dropping implausible %s=%s for sector=%s period=%s",
                            type(self).__name__, metric_name, value_f,
                            sector_code, as_of,
                        )
                        continue

                    records.append({
                        "institution_sector": sector_code,
                        "asset_class": SYNTHETIC_ASSET_CLASS,
                        "period": self._date_to_period_slug(as_of),
                        "metric_name": metric_name,
                        "value": value_f,
                        "as_of_date": as_of,
                        # Audit-trail fields (not required by canonical_columns
                        # but preserved so the scraper can stamp provenance).
                        "_source_sheet": sheet_name,
                        "_source_row": row_idx,
                    })

        if missing_sheets:
            logger.warning(
                "%s: %d of %d configured sector sheets absent: %s",
                type(self).__name__, len(missing_sheets),
                len(self.SECTOR_SHEETS), missing_sheets,
            )

        if not records:
            # Still return a validly-shaped (empty) DataFrame so downstream
            # code doesn't have to special-case the empty case.
            df = pd.DataFrame(columns=self._CANONICAL_COLUMNS)
            self.validate_output(df)
            return df

        df = pd.DataFrame.from_records(records)
        self.validate_output(df)
        return df

    # ---- helpers -----------------------------------------------------------

    def _find_date_row(
        self, rows: list[tuple[Any, ...]],
    ) -> tuple[int, list[tuple[int, date]]]:
        """Return ``(row_index, [(col_index, quarter_end_date), ...])``."""
        return find_date_row(rows, preferred=self.DATE_HEADER_ROW)

    @staticmethod
    def _parse_date_row(row: tuple[Any, ...]) -> list[tuple[int, date]]:
        return parse_date_row(row)

    @classmethod
    def _is_plausible(cls, metric_name: str, value: float) -> bool:
        lo, hi = cls.PLAUSIBILITY.get(metric_name, (0.0, 1.0))
        return lo <= value <= hi

    @staticmethod
    def _date_to_period_slug(d: date) -> str:
        return date_to_period_slug(d)


