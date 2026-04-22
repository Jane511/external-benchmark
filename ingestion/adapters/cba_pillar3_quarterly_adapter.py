"""CBA quarterly Pillar 3 APS 330 supplement adapter (Option A).

The post-2025-01-01 quarterly supplement carries:

- ``CRB(f)(ii)`` — non-performing exposures ($M), specific provision
  balance, and actual losses for the quarter, keyed by portfolio.
- ``EAD & CRWA`` — exposure at default ($M) keyed by asset category
  (which CBA splits across AIRB, FIRB, and standardised approaches).

Neither PD nor LGD is published in this file (that moved to the
half-year / full-year PDF). The adapter computes a per-portfolio NPL
ratio by arithmetic:

    npl_ratio = non_performing_exposures ($M) / total_EAD ($M)

The numerator and denominator are preserved as canonical columns so
MRC can re-derive the ratio without re-opening the workbook — identical
audit-trail convention to the APRA QPEX adapter (Path B).

Portfolio-label matching is label-based (case-insensitive, substring).
Matching AIRB + FIRB + standardised rows for the same portfolio and
summing their EADs is deliberately NOT done here — that is a modelling
choice, not an extraction one. Each approach's EAD is kept separate and
the AIRB row is the primary denominator for IRB-approach asset classes;
if the label only appears in standardised (e.g. "Other assets") the
adapter falls back to that row. Implausible ratios (>10%) are dropped
with a warning.

See ``outputs/cba_pillar3_structure.md`` for the original structural
inspection and column mapping.
"""

from __future__ import annotations

import logging
from datetime import date
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd

from ingestion.adapters.base import AbstractAdapter
from ingestion.adapters.cba_pillar3_pdf_adapter import _derive_cba_period_code

logger = logging.getLogger(__name__)


_NPE_SHEET = "CRB(f)"     # Combined sheet carrying both CRB(f)(i) industry
                          # rows and CRB(f)(ii) portfolio rows. Portfolio
                          # labels like "Residential Mortgage" only appear
                          # in the portfolio section, so label-based row
                          # lookup naturally targets the right section.
_EAD_SHEET = "EAD & CRWA"

# (canonical asset class, matching substring for NPE / EAD rows).
# Matches are case-insensitive; we take the FIRST row in EAD whose label
# contains the pattern — that's the AIRB row when both AIRB and FIRB
# disclose (AIRB is listed first in the sheet layout). "Corporate" in
# CRB(f)(ii) covers both AIRB and FIRB corporate, so we pair it with
# the AIRB corporate EAD row — a conservative choice that favours the
# largest single approach.
_PORTFOLIO_MAP: tuple[tuple[str, tuple[str, ...]], ...] = (
    # CBA's portfolio label for the aggregate corporate row is
    # "Corporate (incl. Large and SME corporate)" in CRB(f)(ii) — the
    # AIRB sub-row in EAD & CRWA is simply "Corporate (incl. SME
    # corporate)". We match on "corporate (incl." to catch both.
    ("corporate_aggregate",   ("corporate (incl.",)),
    ("retail_sme",            ("sme retail",)),
    ("residential_mortgage",  ("residential mortgage",)),
    ("retail_qrr",            ("qualifying revolving retail",)),
    ("retail_other",          ("other retail",)),
    ("specialised_lending",   ("specialised lending",)),
    ("financial_institution", ("financial institution",)),
    ("sovereign",             ("sovereign",)),
)


_NA_TOKENS = {"-", "–", "—", "n/a", "na", ""}


class CbaPillar3QuarterlyAdapter(AbstractAdapter):
    """Compute CBA per-portfolio NPL ratios from the quarterly APS 330 XLSX."""

    _SOURCE_NAME = "cba_pillar3_quarterly"
    _CANONICAL_COLUMNS = [
        "asset_class",
        "metric_name",
        "value",
        "as_of_date",
        "period_code",
        "numerator_value",
        "denominator_value",
        "numerator_sheet",
        "denominator_sheet",
    ]

    PLAUSIBILITY: dict[str, tuple[float, float]] = {
        "npl_ratio": (0.0, 0.10),
    }

    METRIC_NAME = "npl_ratio"
    ARITHMETIC_FORMULA = (
        "npl_ratio = CRB(f)(ii).non_performing_exposures / EAD_CRWA.ead"
    )

    # ---- AbstractAdapter contract ------------------------------------------

    @property
    def source_name(self) -> str:
        return self._SOURCE_NAME

    @property
    def canonical_columns(self) -> list[str]:
        return list(self._CANONICAL_COLUMNS)

    def normalise(
        self,
        file_path: Path,
        *,
        reporting_date: date | str | None = None,
    ) -> pd.DataFrame:
        reporting = _coerce_reporting_date(reporting_date)
        period_code = _derive_cba_period_code(reporting)

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        missing = [s for s in (_NPE_SHEET, _EAD_SHEET) if s not in wb.sheetnames]
        if missing:
            logger.warning(
                "%s: missing required sheets %s (have %s)",
                type(self).__name__, missing, wb.sheetnames[:8],
            )
            df = pd.DataFrame(columns=self._CANONICAL_COLUMNS)
            self.validate_output(df)
            return df

        npe_rows = list(wb[_NPE_SHEET].iter_rows(values_only=True))
        ead_rows = list(wb[_EAD_SHEET].iter_rows(values_only=True))

        npe_by_portfolio = _index_single_value_rows(npe_rows, value_col=1)
        ead_by_portfolio = _index_single_value_rows(ead_rows, value_col=1)

        records: list[dict[str, Any]] = []
        for asset_class, patterns in _PORTFOLIO_MAP:
            numer_label, numer = _first_match(npe_by_portfolio, patterns)
            denom_label, denom = _first_match(ead_by_portfolio, patterns)
            if numer is None or denom is None:
                logger.info(
                    "%s: %s — skipping (npe=%s, ead=%s)",
                    type(self).__name__, asset_class,
                    "found" if numer is not None else "not found",
                    "found" if denom is not None else "not found",
                )
                continue
            if denom == 0.0:
                continue

            ratio = numer / denom
            if not self._is_plausible(self.METRIC_NAME, ratio):
                logger.warning(
                    "%s: dropping implausible %s=%.4f for %s (num=%.2f den=%.2f)",
                    type(self).__name__, self.METRIC_NAME, ratio,
                    asset_class, numer, denom,
                )
                continue

            records.append({
                "asset_class": asset_class,
                "metric_name": self.METRIC_NAME,
                "value": ratio,
                "as_of_date": reporting,
                "period_code": period_code,
                "numerator_value": numer,
                "denominator_value": denom,
                "numerator_sheet": _NPE_SHEET,
                "denominator_sheet": _EAD_SHEET,
            })

        df = pd.DataFrame.from_records(records) if records else \
             pd.DataFrame(columns=self._CANONICAL_COLUMNS)
        self.validate_output(df)
        return df

    @classmethod
    def _is_plausible(cls, metric_name: str, value: float) -> bool:
        lo, hi = cls.PLAUSIBILITY.get(metric_name, (0.0, 1.0))
        return lo <= value <= hi


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _index_single_value_rows(
    rows: list[tuple[Any, ...]], value_col: int,
) -> list[tuple[str, float]]:
    """Return [(label_lower, first_numeric_cell)] for rows with a string label.

    CRB(f)(ii) and EAD & CRWA both use column 0 as the row label. We keep
    every row whose column-0 is a string; the FIRST numeric cell after
    column 0 is the $M value we want (non-performing or EAD respectively).
    """
    out: list[tuple[str, float]] = []
    for row in rows:
        if not row:
            continue
        label = row[0]
        if not isinstance(label, str):
            continue
        lower = label.strip().lower()
        if not lower:
            continue
        # Find first numeric value cell (skip None/blank/NA markers).
        value: float | None = None
        for cell in row[1:]:
            if cell is None:
                continue
            if isinstance(cell, (int, float)):
                value = float(cell)
                break
            if isinstance(cell, str):
                t = cell.strip().replace(",", "")
                if t.lower() in _NA_TOKENS:
                    continue
                try:
                    value = float(t.rstrip("%"))
                    break
                except ValueError:
                    continue
        if value is not None:
            out.append((lower, value))
    return out


def _first_match(
    indexed: list[tuple[str, float]],
    patterns: tuple[str, ...],
) -> tuple[str | None, float | None]:
    """Return the first ``(label, value)`` whose label contains any pattern."""
    for label, value in indexed:
        for p in patterns:
            if p.lower() in label:
                return label, value
    return None, None


def _coerce_reporting_date(value: date | str | None) -> date:
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        return pd.to_datetime(value).date()
    return date.today()
