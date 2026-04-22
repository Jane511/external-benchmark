"""APRA Quarterly Property Exposures adapter (Path B — arithmetic NPL).

Real QPEX workbooks publish dollar-valued commercial and residential
property exposure series across per-sector sheet pairs:

- Commercial sheets:  ``Tab 1a`` (All ADIs), ``Tab 2a`` (Banks),
                      ``Tab 4a`` (Major banks)
- Residential sheets: ``Tab 1b`` / ``Tab 2b`` / ``Tab 4b``

NPL ratios are NOT pre-computed in the residential sheets and are only
partially pre-computed in the commercial sheets. This adapter therefore
produces ``npl_ratio`` by arithmetic — numerator / denominator — and
preserves the two raw dollar values on every row so downstream audit
(MRC) can verify the division without re-opening the workbook.

Row INDICES are NOT identical across the commercial sheets (``Tab 1a``
has a 1-row offset vs ``Tab 2a``/``Tab 4a`` because of an extra footnote).
So the adapter uses **label-based row lookup** — scan column 0 for a
case-insensitive substring pattern that identifies each role. Layout
drift survives as long as APRA keeps the labels.

Time coverage:

- Commercial: APS 220 "Non-performing" rows populate from 2022-Q1 onward
  → 16 quarters × 3 sectors = 48 entries.
- Residential: "Non-performing loans" + "Total credit outstanding"
  populate from 2019-Q1 onward → 28 quarters × 3 sectors = 84 entries.

Expected yield: **132 entries** for a full file; fewer if any sector
sheet is absent.

See ``outputs/apra_workbook_structure.md`` (QPEX deep dive) for the
spec behind these decisions.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd

from ingestion.adapters.apra_helpers import (
    date_to_period_slug,
    find_date_row,
)
from ingestion.adapters.base import AbstractAdapter

logger = logging.getLogger(__name__)


@dataclass(frozen=True)
class _RowMatcher:
    """Declarative spec for finding one semantic row in a sheet.

    ``required`` substrings must ALL appear (case-insensitive) in the
    column-0 label; ``forbidden`` substrings must NONE appear. This
    excludes adjacent sub-breakdown rows (e.g. ``of which: Exposures in
    Australia``) from matching the main total.
    """

    role: str
    required: tuple[str, ...]
    forbidden: tuple[str, ...] = ()


# The label map — edit this if APRA renames a row.
_COMMERCIAL_MATCHERS: tuple[_RowMatcher, ...] = (
    _RowMatcher(
        role="denominator",
        required=("total commercial property exposures",),
        forbidden=("limits", "of which"),
    ),
    _RowMatcher(
        role="numerator",
        required=("non-performing commercial property exposures",),
        forbidden=("of which", "provisions"),
    ),
)

_RESIDENTIAL_MATCHERS: tuple[_RowMatcher, ...] = (
    _RowMatcher(
        role="denominator",
        # APRA's own typo: "oustanding" (not "outstanding"). Match the
        # typo literally; if APRA ever fixes it, add "total credit outstanding".
        required=("total credit o",),  # covers both spellings
        forbidden=("limits", "by purpose", "term loans"),
    ),
    _RowMatcher(
        role="numerator",
        required=("non-performing loans",),
        forbidden=(
            "term loans",        # breakdown rows
            "selected",
            "during the quarter",
            "of which",
            "by lvr",
        ),
    ),
)


_SHEET_TYPE_MATCHERS: dict[str, tuple[_RowMatcher, ...]] = {
    "commercial": _COMMERCIAL_MATCHERS,
    "residential": _RESIDENTIAL_MATCHERS,
}


class ApraQpexAdapter(AbstractAdapter):
    """Extract per-sector NPL ratios for commercial + residential QPEX data."""

    _SOURCE_NAME = "apra_qpex"
    _CANONICAL_COLUMNS = [
        "institution_sector",
        "asset_class",
        "period",
        "metric_name",
        "value",
        "as_of_date",
        "numerator_value",
        "denominator_value",
    ]

    # ---- tunables ---------------------------------------------------------

    SECTOR_SHEET_PAIRS: dict[str, dict[str, str]] = {
        "all_adis":    {"commercial": "Tab 1a", "residential": "Tab 1b"},
        "banks":       {"commercial": "Tab 2a", "residential": "Tab 2b"},
        "major_banks": {"commercial": "Tab 4a", "residential": "Tab 4b"},
    }

    ASSET_CLASS_MAP: dict[str, str] = {
        "commercial":  "commercial_property_investment",
        "residential": "residential_mortgage",
    }

    DATE_HEADER_ROW: int = 3

    PLAUSIBILITY: dict[str, tuple[float, float]] = {
        "npl_ratio": (0.0, 0.10),
    }

    METRIC_NAME = "npl_ratio"
    ARITHMETIC_FORMULA = "npl_ratio = non_performing_dollars / total_exposure_dollars"

    # ---- AbstractAdapter contract -----------------------------------------

    @property
    def source_name(self) -> str:
        return self._SOURCE_NAME

    @property
    def canonical_columns(self) -> list[str]:
        return list(self._CANONICAL_COLUMNS)

    def normalise(self, file_path: Path) -> pd.DataFrame:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        records: list[dict[str, Any]] = []
        missing: list[str] = []

        for sector_code, pair in self.SECTOR_SHEET_PAIRS.items():
            for asset_label, sheet_name in pair.items():
                if sheet_name not in wb.sheetnames:
                    missing.append(sheet_name)
                    logger.warning(
                        "%s: sheet %r not in workbook (sector=%s, asset=%s)",
                        type(self).__name__, sheet_name, sector_code, asset_label,
                    )
                    continue

                rows = list(wb[sheet_name].iter_rows(values_only=True))
                records.extend(self._extract_sheet(
                    rows=rows,
                    sheet_name=sheet_name,
                    sector_code=sector_code,
                    asset_label=asset_label,
                ))

        if missing:
            logger.warning(
                "%s: %d configured sheets absent from workbook: %s",
                type(self).__name__, len(missing), missing,
            )

        if not records:
            df = pd.DataFrame(columns=self._CANONICAL_COLUMNS)
            self.validate_output(df)
            return df

        df = pd.DataFrame.from_records(records)
        self.validate_output(df)
        return df

    # ---- per-sheet extraction ---------------------------------------------

    def _extract_sheet(
        self,
        *,
        rows: list[tuple[Any, ...]],
        sheet_name: str,
        sector_code: str,
        asset_label: str,
    ) -> list[dict[str, Any]]:
        matchers = _SHEET_TYPE_MATCHERS[asset_label]
        matched = self._lookup_rows(rows, matchers)

        if "denominator" not in matched or "numerator" not in matched:
            logger.warning(
                "%s: sheet %r missing required rows (found=%s, needed=[denominator, numerator])",
                type(self).__name__, sheet_name, sorted(matched),
            )
            return []

        try:
            _, dates = find_date_row(rows, preferred=self.DATE_HEADER_ROW)
        except ValueError as exc:
            logger.warning(
                "%s: %s in sheet %r — skipping", type(self).__name__, exc, sheet_name,
            )
            return []

        denom_row_idx, denom_row = matched["denominator"]
        numer_row_idx, numer_row = matched["numerator"]

        asset_class = self.ASSET_CLASS_MAP[asset_label]
        out: list[dict[str, Any]] = []

        for col_idx, as_of in dates:
            denom = _cell_as_float(denom_row, col_idx)
            numer = _cell_as_float(numer_row, col_idx)
            if denom is None or numer is None or denom == 0.0:
                continue

            ratio = numer / denom
            if not self._is_plausible(self.METRIC_NAME, ratio):
                logger.warning(
                    "%s: dropping implausible npl_ratio=%.4f "
                    "(sector=%s asset=%s as_of=%s num=%.2f den=%.2f)",
                    type(self).__name__, ratio, sector_code, asset_class,
                    as_of, numer, denom,
                )
                continue

            out.append({
                "institution_sector": sector_code,
                "asset_class": asset_class,
                "period": date_to_period_slug(as_of),
                "metric_name": self.METRIC_NAME,
                "value": ratio,
                "as_of_date": as_of,
                "numerator_value": numer,
                "denominator_value": denom,
                # Audit trail (preserved beyond canonical columns).
                "_source_sheet": sheet_name,
                "_numerator_row": numer_row_idx,
                "_denominator_row": denom_row_idx,
            })

        return out

    # ---- label-based row lookup -------------------------------------------

    @staticmethod
    def _lookup_rows(
        rows: list[tuple[Any, ...]],
        matchers: tuple[_RowMatcher, ...],
    ) -> dict[str, tuple[int, tuple[Any, ...]]]:
        """Return {role: (row_index, row_tuple)} for each matcher that hits.

        First match wins per role — adjacent sub-breakdown rows (e.g.
        ``Owner-occupied``, ``of which: ...``) are excluded via the
        matchers' ``forbidden`` lists rather than by row position.
        """
        found: dict[str, tuple[int, tuple[Any, ...]]] = {}
        for idx, row in enumerate(rows):
            label = row[0] if row else None
            if not isinstance(label, str):
                continue
            lower = label.strip().lower()
            for m in matchers:
                if m.role in found:
                    continue
                if not all(req in lower for req in m.required):
                    continue
                if any(forb in lower for forb in m.forbidden):
                    continue
                found[m.role] = (idx, row)
        return found

    # ---- small helpers ----------------------------------------------------

    @classmethod
    def _is_plausible(cls, metric_name: str, value: float) -> bool:
        lo, hi = cls.PLAUSIBILITY.get(metric_name, (0.0, 1.0))
        return lo <= value <= hi


def _cell_as_float(row: tuple[Any, ...], col_idx: int) -> float | None:
    if col_idx >= len(row):
        return None
    cell = row[col_idx]
    if cell is None or cell == "":
        return None
    try:
        return float(cell)
    except (TypeError, ValueError):
        return None
