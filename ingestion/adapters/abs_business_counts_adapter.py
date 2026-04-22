"""ABS Counts of Australian Businesses (cat. 8165) adapter.

Reshapes the wide-format multi-year panel in Data Cube 01, Table 1
("Businesses by Industry Division") into a long-format DataFrame of
``(as_of_date, anzsic_division_code, industry, business_count)`` rows
that the existing ``asic_abs_import.py`` pipeline consumes.

Sheet layout (confirmed via inspection, see outputs/asic_abs_structure.md):

- Rows 0-5: title + metadata + column header band.
- Rows 6+: repeating panels, one per fiscal year. Each panel begins with
  a row whose first cell is the FY label (``"2021-22"`` etc.) and has
  no data values; the next ~19 rows are ANZSIC divisions, followed by
  ``"Currently Unknown"`` and ``"All Industries"`` sweep / total rows.
- Column 8 (0-indexed) in each division row is "Operating at end of
  financial year" — the ``business_count`` the engine uses.

Fiscal year end convention: Australian FY ends 30 June, so the label
``"2021-22"`` maps to ``as_of_date = 2022-06-30``.
"""

from __future__ import annotations

import logging
import re
from datetime import date
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd

from ingestion.adapters.base import AbstractAdapter

logger = logging.getLogger(__name__)


# Maps lowercased ABS division labels → ANZSIC 2006 division code.
ANZSIC_DIVISION_CODES: dict[str, str] = {
    "agriculture, forestry and fishing": "A",
    "mining": "B",
    "manufacturing": "C",
    "electricity, gas, water and waste services": "D",
    "construction": "E",
    "wholesale trade": "F",
    "retail trade": "G",
    "accommodation and food services": "H",
    "transport, postal and warehousing": "I",
    "information media and telecommunications": "J",
    "financial and insurance services": "K",
    "rental, hiring and real estate services": "L",
    "professional, scientific and technical services": "M",
    "administrative and support services": "N",
    "public administration and safety": "O",
    "education and training": "P",
    "health care and social assistance": "Q",
    "arts and recreation services": "R",
    "other services": "S",
}


class AbsBusinessCountsAdapter(AbstractAdapter):
    """Reshape ABS 8165 Data Cube 01 Table 1 into long-format business counts."""

    _SOURCE_NAME = "abs_business_counts"
    _CANONICAL_COLUMNS = [
        "as_of_date",
        "anzsic_division_code",
        "industry",
        "business_count",
        "fiscal_year",
        "source_sheet",
        "source_row_label",
    ]

    # ---- tunables ---------------------------------------------------------

    PRIMARY_SHEET = "Table 1"
    BUSINESS_COUNT_COLUMN = 8  # 0-indexed; "Operating at end of financial year"

    # FY section headers appear as "2021-22", "2022-23", etc. — a year-year
    # pair where the second component is two digits.
    FY_HEADER_RE = re.compile(r"^(\d{4})-(\d{2})$")

    EXCLUDE_LABELS: frozenset[str] = frozenset({
        "currently unknown",
        "all industries",
        "total selected industries",
    })

    PLAUSIBILITY: dict[str, tuple[float, float]] = {
        "business_count": (1_000, 10_000_000),
    }

    # ---- AbstractAdapter contract -----------------------------------------

    @property
    def source_name(self) -> str:
        return self._SOURCE_NAME

    @property
    def canonical_columns(self) -> list[str]:
        return list(self._CANONICAL_COLUMNS)

    def normalise(self, file_path: Path) -> pd.DataFrame:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        if self.PRIMARY_SHEET not in wb.sheetnames:
            logger.warning(
                "%s: sheet %r not in workbook (have %s)",
                type(self).__name__, self.PRIMARY_SHEET, wb.sheetnames[:8],
            )
            df = pd.DataFrame(columns=self._CANONICAL_COLUMNS)
            self.validate_output(df)
            return df

        rows = list(wb[self.PRIMARY_SHEET].iter_rows(values_only=True))

        records: list[dict[str, Any]] = []
        current_fy_end: date | None = None

        for row in rows:
            if not row:
                continue
            first = row[0]
            if first is None:
                continue
            first_text = str(first).strip()
            if not first_text:
                continue

            fy_match = self.FY_HEADER_RE.match(first_text)
            if fy_match:
                # "2021-22" → FY ending 30 June 2022
                current_fy_end = date(int(fy_match.group(1)) + 1, 6, 30)
                continue

            if current_fy_end is None:
                continue

            first_lower = first_text.lower()
            if first_lower in self.EXCLUDE_LABELS:
                continue

            code = ANZSIC_DIVISION_CODES.get(first_lower)
            if code is None:
                # Unknown row label — likely a sub-heading or note. Skip
                # silently rather than warn because these are common.
                continue

            if self.BUSINESS_COUNT_COLUMN >= len(row):
                continue
            raw = row[self.BUSINESS_COUNT_COLUMN]
            if raw is None:
                continue
            try:
                count = int(float(raw))
            except (TypeError, ValueError):
                continue

            if not self._is_plausible("business_count", count):
                logger.warning(
                    "%s: dropping implausible business_count=%d for %s FY%s",
                    type(self).__name__, count, first_text, current_fy_end.year,
                )
                continue

            records.append({
                "as_of_date": current_fy_end,
                "anzsic_division_code": code,
                "industry": first_text,
                "business_count": count,
                "fiscal_year": f"FY{current_fy_end.year}",
                "source_sheet": self.PRIMARY_SHEET,
                "source_row_label": first_text,
            })

        df = pd.DataFrame.from_records(records) if records else \
             pd.DataFrame(columns=self._CANONICAL_COLUMNS)
        self.validate_output(df)
        return df

    @classmethod
    def _is_plausible(cls, metric_name: str, value: float) -> bool:
        lo, hi = cls.PLAUSIBILITY.get(metric_name, (0.0, float("inf")))
        return lo <= value <= hi
